import asyncio
import io
import json
import logging
import mimetypes
import os
import re
import tempfile
import uuid
from contextlib import asynccontextmanager
from typing import AsyncIterator

import asyncpg
import httpx
import websockets
from authlib.integrations.starlette_client import OAuth
from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, Query, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, RedirectResponse
from jose import JWTError, jwt
from piper.voice import PiperVoice
from pydantic import BaseModel
from starlette.config import Config
from starlette.middleware.sessions import SessionMiddleware
from starlette.responses import StreamingResponse

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("meeting-analyzer")

WHISPER_URL = os.getenv("WHISPER_URL", "http://whisper:9000")
WHISPER_LIVE_URL = os.getenv("WHISPER_LIVE_URL", "ws://whisper-live:9090")
PIPER_TTS_URL = os.getenv("PIPER_TTS_URL", "http://piper-tts:5000")
CLAUDE_API_KEY = os.getenv("CLAUDE_API_KEY", "")
PIPER_MODEL_PATH = os.getenv("PIPER_MODEL_PATH", "/models/en_US-lessac-medium.onnx")
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://analyzer:analyzer-pass-2026@analyzer-postgres:5432/analyzer",
)
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "minio.whisper.svc.cluster.local:9000")
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "mp4-reader")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "minio-mp4reader")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "minio-mp4reader-secret-2026")

# Keycloak OAuth2/OIDC Configuration
KEYCLOAK_URL = os.getenv("KEYCLOAK_URL", "http://keycloak.keycloak.svc.cluster.local:8080")
KEYCLOAK_REALM = os.getenv("KEYCLOAK_REALM", "mycluster")
KEYCLOAK_CLIENT_ID = os.getenv("KEYCLOAK_CLIENT_ID", "meeting-analyzer")
KEYCLOAK_CLIENT_SECRET = os.getenv("KEYCLOAK_CLIENT_SECRET", "")
KEYCLOAK_CALLBACK_URL = os.getenv("KEYCLOAK_CALLBACK_URL", "http://localhost:30903/auth/callback")
SESSION_SECRET_KEY = os.getenv("SESSION_SECRET_KEY", "")
APP_PATH_PREFIX = os.getenv("APP_PATH_PREFIX", "")

# LiveKit config (deployed in 'robert' namespace, accessible via Tailscale)
LIVEKIT_WS_URL = os.getenv("LIVEKIT_WS_URL", "wss://jetson-orin.desmana-truck.ts.net")
LIVEKIT_INTERNAL_URL = os.getenv("LIVEKIT_INTERNAL_URL", "http://livekit-service.robert.svc.cluster.local:7880")
LIVEKIT_API_KEY = os.getenv("LIVEKIT_API_KEY", "devkey")
LIVEKIT_API_SECRET = os.getenv("LIVEKIT_API_SECRET", "secret123456789")

# Team Chat Widget (Matrix-based agent chat)
TEAM_CHAT_HOMESERVER = os.getenv("TEAM_CHAT_HOMESERVER", "")
TEAM_CHAT_TOKEN = os.getenv("TEAM_CHAT_TOKEN", "")
TEAM_CHAT_USER_ID = os.getenv("TEAM_CHAT_USER_ID", "")
TEAM_CHAT_ROOM_ID = os.getenv("TEAM_CHAT_ROOM_ID", "")

db_pool: asyncpg.Pool | None = None
piper_voice: PiperVoice | None = None
minio_client = None  # aiobotocore S3 client session


class WorkspaceCreate(BaseModel):
    name: str


class ChatRequest(BaseModel):
    meeting_ids: list[int]
    include_transcripts: list[int] = []
    include_document_ids: list[int] = []
    question: str


class TTSRequest(BaseModel):
    text: str


class AnalyzeTextRequest(BaseModel):
    text: str
    workspace_id: int | None = None


async def init_db():
    global db_pool
    db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)
    async with db_pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS meetings (
                id SERIAL PRIMARY KEY,
                title TEXT NOT NULL,
                filename TEXT,
                date TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                transcript TEXT,
                summary TEXT,
                action_items JSONB DEFAULT '[]'::jsonb,
                email_body TEXT,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS workspaces (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        await conn.execute("""
            ALTER TABLE meetings
            ADD COLUMN IF NOT EXISTS workspace_id INTEGER REFERENCES workspaces(id) ON DELETE SET NULL
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS documents (
                id SERIAL PRIMARY KEY,
                workspace_id INTEGER NOT NULL REFERENCES workspaces(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                object_key TEXT NOT NULL UNIQUE,
                file_size BIGINT NOT NULL,
                mime_type TEXT NOT NULL DEFAULT 'application/octet-stream',
                extracted_text TEXT,
                uploaded_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
        """)
        await conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_documents_workspace ON documents(workspace_id)
        """)


async def _try_connect_minio() -> bool:
    """Attempt a single MinIO connection. Returns True on success, False on failure."""
    global minio_client
    try:
        import aiobotocore.session as aio_session
        session = aio_session.get_session()
        # Verify connectivity and create bucket if missing
        async with session.create_client(
            "s3",
            endpoint_url=f"http://{MINIO_ENDPOINT}",
            aws_access_key_id=MINIO_ACCESS_KEY,
            aws_secret_access_key=MINIO_SECRET_KEY,
            region_name="us-east-1",
        ) as client:
            try:
                await client.head_bucket(Bucket=MINIO_BUCKET)
                logger.info("MinIO bucket '%s' ready", MINIO_BUCKET)
            except Exception:
                await client.create_bucket(Bucket=MINIO_BUCKET)
                logger.info("MinIO bucket '%s' created", MINIO_BUCKET)
        # Store session only after successful connectivity check
        minio_client = session
        return True
    except Exception as e:
        logger.warning("MinIO init failed (documents will be unavailable): %s", e)
        minio_client = None
        return False


async def _init_minio():
    """Initialize MinIO client, with background retry if the first attempt fails."""
    if not await _try_connect_minio():
        asyncio.create_task(_minio_retry_loop())


async def _minio_retry_loop():
    """Retry connecting to MinIO every 30 seconds until successful."""
    retry_interval = 30
    while True:
        await asyncio.sleep(retry_interval)
        logger.info("Retrying MinIO connection...")
        if await _try_connect_minio():
            logger.info("MinIO connection established on retry.")
            return


@asynccontextmanager
async def lifespan(app):
    global piper_voice
    await init_db()
    await _init_minio()
    # Pre-warm Keycloak OIDC metadata with internal endpoint URLs.
    # authlib fetches metadata from server_metadata_url and caches it, overwriting
    # any pre-configured values. We populate the cache here so the fetched
    # jwks_uri (which uses the HTTPS frontend URL the pod can't reach) never gets used.
    try:
        import httpx, time
        async with httpx.AsyncClient() as _hc:
            _meta_resp = await _hc.get(
                f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}/.well-known/openid-configuration",
                timeout=10,
            )
            _meta_resp.raise_for_status()
            _metadata = _meta_resp.json()
        _internal_base = f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}/protocol/openid-connect"
        _metadata["jwks_uri"] = f"{_internal_base}/certs"
        _metadata["token_endpoint"] = f"{_internal_base}/token"
        _metadata["userinfo_endpoint"] = f"{_internal_base}/userinfo"
        _metadata["_loaded_at"] = time.time()
        oauth.keycloak.server_metadata.update(_metadata)
        logger.info("Keycloak OIDC metadata pre-warmed with internal endpoints")
    except Exception as _e:
        logger.warning("Failed to pre-warm Keycloak metadata: %s", _e)
    if os.path.exists(PIPER_MODEL_PATH):
        piper_voice = PiperVoice.load(PIPER_MODEL_PATH)
        logger.info("Piper TTS loaded: sample_rate=%d", piper_voice.config.sample_rate)
        # Warm up ONNX runtime to avoid ~500ms cold-start on first request
        await asyncio.to_thread(lambda: list(piper_voice.synthesize("Hello.")))
        logger.info("Piper ONNX runtime warmed up")
    else:
        logger.warning("Piper model not found at %s — streaming TTS disabled", PIPER_MODEL_PATH)
    yield
    if db_pool:
        await db_pool.close()


app = FastAPI(title="Meeting Analyzer", lifespan=lifespan)

# Add session middleware for OAuth
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET_KEY or os.urandom(32).hex(),
    session_cookie="meeting_analyzer_session",
    max_age=3600 * 24,  # 24 hours
)

# Configure OAuth with Keycloak
oauth = OAuth()
oauth.register(
    name="keycloak",
    client_id=KEYCLOAK_CLIENT_ID,
    client_secret=KEYCLOAK_CLIENT_SECRET,
    server_metadata_url=f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}/.well-known/openid-configuration",
    # Override server-side endpoints to use internal cluster URL.
    # The OIDC discovery doc returns HTTPS frontend URLs (for the browser),
    # but the pod cannot reach those — it uses the internal service instead.
    access_token_url=f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}/protocol/openid-connect/token",
    jwks_uri=f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}/protocol/openid-connect/certs",
    client_kwargs={
        "scope": "openid email profile",
        "code_challenge_method": "S256",
    },
)

ANALYSIS_PROMPT = """\
You are a meeting analyst. Given the following transcript, produce:

1. A short title for this meeting (5-8 words max)
2. A concise summary (2-4 sentences)
3. A list of action items with assignee if mentioned
4. A professional email body summarizing this meeting that could be sent to attendees. The email should include a brief greeting, the summary, the action items as a bulleted list, and a professional sign-off. Do not include a subject line or To/From headers.

Format your response as JSON with exactly these keys:
- "title": string
- "summary": string
- "action_items": list of strings
- "email_body": string

Respond with ONLY valid JSON, no markdown fences.

<transcript>
{transcript}
</transcript>"""

CHAT_SYSTEM_PROMPT = (
    "You are a helpful meeting analyst. You have access to meeting data provided "
    "in XML format. Answer the user's question based on the meeting content. Be "
    "specific and reference meeting details when relevant. If the information is not "
    "in the provided meetings, say so."
)

HTML_PAGE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Meeting Analyzer</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=Literata:opsz,wght@7..72,400;7..72,500&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {
    --bg: #ffffff; --bg-warm: #f8f6f3;
    --surface: #f2eeea; --surface-raised: #ebe6e0;
    --border: #ddd5cb; --border-light: #c9bfb3;
    --text: #2c2419; --text-secondary: #6b5d4f; --text-dim: #9a8d7f;
    --accent: #a07040; --accent-glow: #b8844e; --accent-dim: #c49a68;
    --red: #b84040; --red-bg: #fdf0f0; --green: #4a8a4a;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Literata', Georgia, serif; background: var(--bg);
    color: var(--text); min-height: 100vh; -webkit-font-smoothing: antialiased;
  }
  body::before {
    content: ''; position: fixed; inset: 0; z-index: 0; pointer-events: none; opacity: 0.02;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
    background-size: 256px 256px;
  }
  .page { position: relative; z-index: 1; max-width: 720px; margin: 0 auto; padding: 4rem 1.5rem 6rem; }

  header { margin-bottom: 2.5rem; }
  .logo { display: flex; align-items: center; gap: 0.6rem; margin-bottom: 0.75rem; }
  .logo-icon {
    width: 32px; height: 32px; border-radius: 6px; background: var(--accent);
    display: flex; align-items: center; justify-content: center;
  }
  .logo-icon svg { width: 18px; height: 18px; }
  h1 {
    font-family: 'DM Serif Display', Georgia, serif;
    font-size: 2.4rem; font-weight: 400; letter-spacing: -0.02em; line-height: 1.1;
  }
  .tagline { font-size: 1.05rem; color: var(--text-secondary); margin-top: 0.5rem; line-height: 1.5; }
  /* User area */
  .user-area {
    position: absolute; top: 1rem; right: 1.5rem; display: flex; align-items: center; gap: 0.75rem;
    font-family: 'JetBrains Mono', monospace; font-size: 0.78rem;
  }
  .user-name { color: var(--text-secondary); }
  .logout-btn {
    background: none; border: 1px solid var(--border); color: var(--text-dim);
    padding: 0.3rem 0.65rem; border-radius: 6px; cursor: pointer; transition: all 0.2s;
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem; text-transform: uppercase;
  }
  .logout-btn:hover { border-color: var(--red); color: var(--red); }


  /* Section headers */
  .section-header {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 1rem; padding-bottom: 0.5rem; border-bottom: 1px solid var(--border);
  }
  .section-title {
    font-family: 'JetBrains Mono', monospace; font-size: 0.75rem;
    text-transform: uppercase; letter-spacing: 0.08em; color: var(--text-dim);
  }

  /* Buttons */
  .accent-btn {
    background: none; border: 1px solid var(--accent-dim); color: var(--accent);
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem;
    padding: 0.35rem 0.9rem; border-radius: 6px; cursor: pointer; transition: all 0.2s;
  }
  .accent-btn:hover { background: var(--accent-dim); color: var(--bg); }

  /* New workspace form */
  .new-ws-form {
    display: flex; gap: 0.5rem; margin-bottom: 1rem; align-items: center;
  }
  .new-ws-input {
    flex: 1; background: var(--surface); border: 1px solid var(--border);
    border-radius: 8px; color: var(--text); font-family: 'Literata', serif;
    font-size: 0.9rem; padding: 0.5rem 0.75rem; outline: none;
  }
  .new-ws-input:focus { border-color: var(--accent-dim); }
  .new-ws-submit {
    background: var(--accent); border: none; color: var(--bg); border-radius: 8px;
    font-family: 'JetBrains Mono', monospace; font-size: 0.75rem;
    padding: 0.5rem 1rem; cursor: pointer; transition: all 0.2s;
  }
  .new-ws-submit:hover { background: var(--accent-glow); }
  .new-ws-cancel {
    background: none; border: 1px solid var(--border); color: var(--text-dim);
    border-radius: 8px; font-family: 'JetBrains Mono', monospace; font-size: 0.75rem;
    padding: 0.5rem 0.75rem; cursor: pointer; transition: all 0.2s;
  }
  .new-ws-cancel:hover { color: var(--text-secondary); border-color: var(--border-light); }

  /* Workspace grid */
  .ws-grid {
    display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
    gap: 0.75rem; margin-bottom: 2.5rem;
  }
  .ws-card {
    background: var(--surface); border: 1px solid var(--border); border-radius: 12px;
    padding: 1.25rem; cursor: pointer; transition: all 0.2s;
  }
  .ws-card:hover {
    border-color: var(--border-light); background: var(--surface-raised);
    transform: translateY(-1px);
  }
  .ws-card-name {
    font-family: 'DM Serif Display', Georgia, serif;
    font-size: 1.05rem; color: var(--text); margin-bottom: 0.3rem;
  }
  .ws-card-count {
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem; color: var(--text-dim);
  }
  .empty-state { color: var(--text-dim); text-align: center; padding: 2rem; font-size: 0.9rem; }

  /* Workspace nav */
  .ws-nav {
    display: flex; align-items: center; gap: 1rem; margin-bottom: 1.5rem;
  }
  .back-btn {
    background: none; border: 1px solid var(--border); color: var(--text-secondary);
    font-family: 'Literata', serif; font-size: 0.85rem;
    padding: 0.35rem 0.8rem; border-radius: 8px; cursor: pointer; transition: all 0.2s;
    display: flex; align-items: center; gap: 0.3rem;
  }
  .back-btn:hover { border-color: var(--border-light); color: var(--text); }
  .ws-name {
    font-family: 'DM Serif Display', Georgia, serif;
    font-size: 1.5rem; font-weight: 400; flex: 1;
  }
  .del-ws-btn {
    background: none; border: 1px solid rgba(196,112,112,0.25); color: var(--red);
    font-family: 'JetBrains Mono', monospace; font-size: 0.7rem;
    padding: 0.3rem 0.7rem; border-radius: 6px; cursor: pointer; transition: all 0.2s;
  }
  .del-ws-btn:hover { background: var(--red-bg); }

  /* Tabs */
  .tabs { display: flex; gap: 0; margin-bottom: 2rem; border-bottom: 1px solid var(--border); }
  .tab {
    padding: 0.6rem 1.25rem; font-family: 'JetBrains Mono', monospace;
    font-size: 0.78rem; text-transform: uppercase; letter-spacing: 0.08em;
    color: var(--text-dim); cursor: pointer; border-bottom: 2px solid transparent;
    margin-bottom: -1px; transition: all 0.2s; background: none;
    border-top: none; border-left: none; border-right: none;
  }
  .tab:hover { color: var(--text-secondary); }
  .tab.active { color: var(--accent); border-bottom-color: var(--accent); }
  .tab-content { display: none; }
  .tab-content.active { display: block; }

  /* Upload zone */
  .upload-zone {
    border: 1.5px dashed var(--border-light); border-radius: 16px;
    padding: 3.5rem 2rem; text-align: center; cursor: pointer;
    transition: all 0.35s cubic-bezier(0.4, 0, 0.2, 1);
    background: var(--surface); position: relative; overflow: hidden;
  }
  .upload-zone::after {
    content: ''; position: absolute; inset: 0; border-radius: 16px;
    background: radial-gradient(ellipse at 50% 0%, rgba(160,112,64,0.06) 0%, transparent 70%);
    pointer-events: none; transition: opacity 0.35s;
  }
  .upload-zone:hover, .upload-zone.dragover {
    border-color: var(--accent-dim); background: var(--surface-raised);
    transform: translateY(-2px);
    box-shadow: 0 8px 32px rgba(0,0,0,0.08), 0 0 0 1px rgba(160,112,64,0.15);
  }
  .upload-zone:hover::after, .upload-zone.dragover::after { opacity: 2; }
  .upload-icon {
    width: 56px; height: 56px; margin: 0 auto 1.25rem; border-radius: 50%;
    background: var(--bg); display: flex; align-items: center; justify-content: center;
    border: 1px solid var(--border); transition: border-color 0.35s, background 0.35s;
  }
  .upload-zone:hover .upload-icon { border-color: var(--accent-dim); background: var(--bg-warm); }
  .upload-icon svg { width: 24px; height: 24px; color: var(--accent); }
  .upload-title {
    font-family: 'DM Serif Display', Georgia, serif;
    font-size: 1.15rem; color: var(--text); margin-bottom: 0.4rem;
  }
  .upload-hint { font-size: 0.88rem; color: var(--text-dim); }
  .upload-zone input { display: none; }
  .file-chip {
    display: none; margin-top: 1rem;
    font-family: 'JetBrains Mono', monospace; font-size: 0.8rem;
    color: var(--accent); background: var(--surface);
    border: 1px solid var(--border); border-radius: 100px;
    padding: 0.4rem 1rem; width: fit-content;
  }
  .file-chip.active { display: inline-flex; align-items: center; gap: 0.5rem; }
  .file-chip::before { content: ''; width: 6px; height: 6px; border-radius: 50%; background: var(--accent); flex-shrink: 0; }

  /* Pipeline */
  .pipeline { display: none; margin: 2rem 0; }
  .pipeline.active { display: block; }
  .pipe-steps { display: flex; flex-direction: column; gap: 0; }
  .pipe-step {
    display: flex; align-items: flex-start; gap: 1rem;
    padding: 0.75rem 0; position: relative; opacity: 0.3; transition: opacity 0.4s;
  }
  .pipe-step.current, .pipe-step.done { opacity: 1; }
  .pipe-step .pipe-line { width: 24px; display: flex; flex-direction: column; align-items: center; flex-shrink: 0; }
  .pipe-dot {
    width: 10px; height: 10px; border-radius: 50%; border: 1.5px solid var(--border-light);
    background: var(--bg); transition: all 0.3s; flex-shrink: 0;
  }
  .pipe-step.current .pipe-dot {
    border-color: var(--accent); background: var(--accent);
    box-shadow: 0 0 12px rgba(212,165,116,0.4);
    animation: pulse-dot 1.8s ease-in-out infinite;
  }
  .pipe-step.done .pipe-dot { border-color: var(--green); background: var(--green); }
  .pipe-connector { width: 1.5px; background: var(--border); flex-grow: 1; min-height: 16px; }
  .pipe-step.done .pipe-connector { background: var(--green); }
  .pipe-step:last-child .pipe-connector { display: none; }
  .pipe-label { font-size: 0.9rem; color: var(--text-secondary); line-height: 1.4; }
  .pipe-step.current .pipe-label { color: var(--accent); font-weight: 500; animation: pulse-label 1.4s ease-in-out infinite; }
  .pipe-step.done .pipe-label { color: var(--text-dim); text-decoration: line-through; text-decoration-color: var(--border); }
  .pipe-elapsed {
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem;
    color: var(--accent); opacity: 0.7; margin-left: 0.5rem;
  }
  @keyframes pulse-dot {
    0%, 100% { box-shadow: 0 0 8px rgba(160,112,64,0.3); }
    50% { box-shadow: 0 0 16px rgba(160,112,64,0.5); }
  }
  @keyframes pulse-label {
    0%, 100% { opacity: 1; color: var(--accent); }
    50% { opacity: 0.35; color: var(--text-dim); }
  }

  /* Error */
  .error-box {
    display: none; margin: 1.5rem 0; padding: 1rem 1.25rem;
    background: var(--red-bg); border: 1px solid rgba(196,112,112,0.25);
    border-radius: 10px; color: var(--red); font-size: 0.92rem; line-height: 1.5;
  }
  .error-box.active { display: flex; align-items: flex-start; gap: 0.75rem; }
  .error-box svg { flex-shrink: 0; margin-top: 2px; }

  /* Results */
  .results { display: none; }
  .results.active { display: block; animation: fadeUp 0.5s cubic-bezier(0.4, 0, 0.2, 1); }
  @keyframes fadeUp {
    from { opacity: 0; transform: translateY(16px); }
    to { opacity: 1; transform: translateY(0); }
  }
  .result-section {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 12px; padding: 1.5rem; margin-bottom: 1rem;
    animation: fadeUp 0.5s cubic-bezier(0.4, 0, 0.2, 1) backwards;
  }
  .result-section:nth-child(1) { animation-delay: 0s; }
  .result-section:nth-child(2) { animation-delay: 0.1s; }
  .result-section:nth-child(3) { animation-delay: 0.2s; }
  .result-section:nth-child(4) { animation-delay: 0.3s; }
  .result-section .label {
    font-family: 'JetBrains Mono', monospace; font-size: 0.7rem;
    text-transform: uppercase; letter-spacing: 0.1em;
    color: var(--accent); margin-bottom: 0.85rem;
    display: flex; align-items: center; gap: 0.5rem;
  }
  .label-line { flex: 1; height: 1px; background: var(--border); }
  .result-section p { font-size: 1rem; line-height: 1.75; }
  .result-section ul { list-style: none; padding: 0; }
  .result-section li {
    padding: 0.6rem 0; border-bottom: 1px solid var(--border);
    font-size: 0.95rem; line-height: 1.6;
    display: flex; align-items: flex-start; gap: 0.75rem;
  }
  .result-section li:last-child { border-bottom: none; padding-bottom: 0; }
  .result-section li::before {
    content: ''; width: 5px; height: 5px; border-radius: 50%;
    background: var(--accent); flex-shrink: 0; margin-top: 0.55rem;
  }
  .transcript-box, .email-box {
    max-height: 360px; overflow-y: auto;
    font-family: 'JetBrains Mono', monospace; font-size: 0.82rem;
    line-height: 1.8; color: var(--text-secondary);
    background: var(--bg); border: 1px solid var(--border);
    border-radius: 8px; padding: 1.25rem; white-space: pre-wrap;
  }
  .email-box { font-family: 'Literata', Georgia, serif; font-size: 0.92rem; line-height: 1.7; max-height: 400px; }
  .transcript-box::-webkit-scrollbar, .email-box::-webkit-scrollbar { width: 6px; }
  .transcript-box::-webkit-scrollbar-track, .email-box::-webkit-scrollbar-track { background: transparent; }
  .transcript-box::-webkit-scrollbar-thumb, .email-box::-webkit-scrollbar-thumb { background: var(--border-light); border-radius: 3px; }

  .copy-btn {
    margin-top: 0.6rem; float: right;
    background: none; border: 1px solid var(--border); color: var(--text-dim);
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem;
    padding: 0.3rem 0.8rem; border-radius: 6px; cursor: pointer;
    transition: all 0.2s; text-transform: uppercase; letter-spacing: 0.05em;
  }
  .copy-btn:hover { border-color: var(--accent-dim); color: var(--accent); }

  .tts-btn {
    margin-top: 0.6rem; float: right; margin-right: 0.5rem;
    background: none; border: 1px solid var(--border); color: var(--text-dim);
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem;
    padding: 0.3rem 0.8rem; border-radius: 6px; cursor: pointer;
    transition: all 0.2s; text-transform: uppercase; letter-spacing: 0.05em;
  }
  .tts-btn:hover { border-color: var(--accent-dim); color: var(--accent); }
  .tts-btn.playing { border-color: var(--accent); color: var(--accent); background: rgba(160,112,64,0.08); }

  .reset-row { text-align: center; margin-top: 1.5rem; clear: both; }
  .reset-btn {
    background: none; border: 1px solid var(--border); color: var(--text-secondary);
    font-family: 'Literata', serif; font-size: 0.88rem;
    padding: 0.5rem 1.5rem; border-radius: 100px; cursor: pointer; transition: all 0.25s;
  }
  .reset-btn:hover { border-color: var(--accent-dim); color: var(--accent); }

  /* Meeting cards */
  .meeting-card {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 12px; padding: 1.25rem; margin-bottom: 0.75rem;
    cursor: pointer; transition: all 0.2s;
  }
  .meeting-card:hover {
    border-color: var(--border-light); background: var(--surface-raised);
    transform: translateY(-1px);
  }
  .mc-title {
    font-family: 'DM Serif Display', Georgia, serif;
    font-size: 1.05rem; color: var(--text); margin-bottom: 0.3rem;
  }
  .mc-meta {
    font-family: 'JetBrains Mono', monospace; font-size: 0.72rem;
    color: var(--text-dim); margin-bottom: 0.5rem;
  }
  .mc-summary {
    font-size: 0.9rem; color: var(--text-secondary); line-height: 1.6;
    display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden;
  }

  /* Detail modal */
  .detail-overlay {
    display: none; position: fixed; inset: 0; z-index: 100;
    background: rgba(0,0,0,0.35); backdrop-filter: blur(4px);
  }
  .detail-overlay.active { display: flex; align-items: flex-start; justify-content: center; padding: 4rem 1rem; overflow-y: auto; }
  .detail-panel {
    background: var(--bg-warm); border: 1px solid var(--border);
    border-radius: 16px; max-width: 720px; width: 100%;
    padding: 2rem; animation: fadeUp 0.3s ease;
  }
  .detail-close {
    float: right; background: none; border: none; color: var(--text-dim);
    cursor: pointer; font-size: 1.2rem; padding: 0.25rem; transition: color 0.2s;
  }
  .detail-close:hover { color: var(--text); }
  .detail-title {
    font-family: 'DM Serif Display', Georgia, serif;
    font-size: 1.5rem; margin-bottom: 0.25rem; padding-right: 2rem;
  }
  .detail-date {
    font-family: 'JetBrains Mono', monospace; font-size: 0.75rem;
    color: var(--text-dim); margin-bottom: 1.5rem;
  }

  /* Chat */
  .chat-container { display: flex; flex-direction: column; }
  .meeting-selector {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 12px; padding: 1rem; margin-bottom: 1rem;
  }
  .ms-header {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 0.6rem;
  }
  .ms-list { display: flex; flex-direction: column; gap: 0.25rem; max-height: 180px; overflow-y: auto; }
  .ms-list::-webkit-scrollbar { width: 4px; }
  .ms-list::-webkit-scrollbar-thumb { background: var(--border-light); border-radius: 2px; }
  .ms-item {
    display: flex; align-items: center; gap: 0.6rem; padding: 0.35rem 0.25rem;
    border-radius: 6px; font-size: 0.85rem;
  }
  .ms-item:hover { background: var(--surface-raised); }
  .ms-item input[type="checkbox"] { accent-color: var(--accent); cursor: pointer; flex-shrink: 0; }
  .ms-item-title { flex: 1; color: var(--text); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .ms-item-date {
    font-family: 'JetBrains Mono', monospace; font-size: 0.68rem;
    color: var(--text-dim); flex-shrink: 0;
  }
  .ms-transcript-label {
    font-family: 'JetBrains Mono', monospace; font-size: 0.62rem; color: var(--text-dim);
    display: flex; align-items: center; gap: 0.25rem; cursor: pointer; flex-shrink: 0;
    padding: 0.15rem 0.4rem; border: 1px solid var(--border); border-radius: 4px;
    transition: all 0.2s;
  }
  .ms-transcript-label:has(input:checked) { border-color: var(--accent-dim); color: var(--accent); }
  .ms-transcript-label input { display: none; }

  .chat-messages {
    min-height: 300px; max-height: 500px; overflow-y: auto;
    display: flex; flex-direction: column; gap: 0.75rem;
    padding: 1rem 0; margin-bottom: 0.75rem;
  }
  .chat-messages::-webkit-scrollbar { width: 5px; }
  .chat-messages::-webkit-scrollbar-thumb { background: var(--border-light); border-radius: 3px; }
  .chat-msg {
    max-width: 85%; padding: 0.75rem 1rem; border-radius: 12px;
    font-size: 0.92rem; line-height: 1.65; white-space: pre-wrap; word-break: break-word;
  }
  .chat-msg.user {
    align-self: flex-end; background: var(--accent-dim); color: var(--text);
    border-bottom-right-radius: 4px;
  }
  .chat-msg.assistant {
    align-self: flex-start; background: var(--surface); border: 1px solid var(--border);
    color: var(--text); border-bottom-left-radius: 4px;
  }
  .chat-msg.assistant.streaming .msg-content::after {
    content: '\\25CA'; margin-left: 2px; animation: blink-cursor 0.8s infinite;
    color: var(--accent);
  }
  @keyframes blink-cursor { 0%, 100% { opacity: 1; } 50% { opacity: 0; } }
  .chat-msg.warning {
    align-self: center; background: none; border: 1px solid rgba(160,112,64,0.3);
    color: var(--accent); font-size: 0.8rem; padding: 0.4rem 0.8rem;
    border-radius: 100px; max-width: 100%;
    font-family: 'JetBrains Mono', monospace;
  }
  .chat-empty {
    flex: 1; display: flex; align-items: center; justify-content: center;
    color: var(--text-dim); font-size: 0.9rem; min-height: 200px;
  }

  .chat-input-area {
    display: flex; gap: 0.5rem; align-items: flex-end;
    padding-top: 0.75rem; border-top: 1px solid var(--border);
  }
  .chat-input-area textarea {
    flex: 1; background: var(--surface); border: 1px solid var(--border);
    border-radius: 8px; color: var(--text); font-family: 'Literata', Georgia, serif;
    font-size: 0.9rem; padding: 0.6rem 0.8rem; resize: none; line-height: 1.5;
    min-height: 42px; max-height: 120px; outline: none;
  }
  .chat-input-area textarea:focus { border-color: var(--accent-dim); }
  .chat-input-area textarea::placeholder { color: var(--text-dim); }
  .chat-send-btn {
    background: var(--accent); border: none; border-radius: 8px; color: var(--bg);
    cursor: pointer; padding: 0.6rem 0.7rem; transition: all 0.2s;
    display: flex; align-items: center; justify-content: center; flex-shrink: 0;
  }
  .chat-send-btn:hover { background: var(--accent-glow); }
  .chat-send-btn:disabled { opacity: 0.35; cursor: not-allowed; }
  .chat-send-btn svg { width: 18px; height: 18px; }

  .voice-toggle {
    background: none; border: 1px solid var(--border); border-radius: 8px; color: var(--text-dim);
    cursor: pointer; padding: 0.6rem 0.7rem; transition: all 0.2s;
    display: flex; align-items: center; justify-content: center; flex-shrink: 0;
  }
  .voice-toggle:hover { border-color: var(--accent-dim); color: var(--accent); }
  .voice-toggle.active { border-color: var(--accent); color: var(--accent); background: rgba(160,112,64,0.08); }
  .voice-toggle svg { width: 18px; height: 18px; }

  /* Live Transcribe */
  .live-container { display: flex; flex-direction: column; align-items: center; }
  .mic-btn {
    width: 72px; height: 72px; border-radius: 50%; border: 2px solid var(--border);
    background: var(--surface); cursor: pointer; transition: all 0.3s;
    display: flex; align-items: center; justify-content: center; margin: 1.5rem 0;
  }
  .mic-btn svg { width: 28px; height: 28px; color: var(--text-secondary); transition: color 0.3s; }
  .mic-btn:hover { border-color: var(--accent-dim); background: var(--surface-raised); }
  .mic-btn:hover svg { color: var(--accent); }
  .mic-btn.recording {
    border-color: var(--red); background: var(--red-bg);
    animation: mic-pulse 1.5s ease-in-out infinite;
  }
  .mic-btn.recording svg { color: var(--red); }
  @keyframes mic-pulse {
    0%, 100% { box-shadow: 0 0 0 0 rgba(184,64,64,0.3); }
    50% { box-shadow: 0 0 0 12px rgba(184,64,64,0); }
  }
  .mic-label {
    font-family: 'JetBrains Mono', monospace; font-size: 0.75rem;
    color: var(--text-dim); text-transform: uppercase; letter-spacing: 0.08em;
  }
  .live-transcript {
    width: 100%; margin-top: 1.5rem; background: var(--surface);
    border: 1px solid var(--border); border-radius: 12px;
    padding: 1.25rem; min-height: 200px; max-height: 400px;
    overflow-y: auto; font-size: 0.95rem; line-height: 1.75;
    color: var(--text); white-space: pre-wrap;
  }
  .live-transcript::-webkit-scrollbar { width: 6px; }
  .live-transcript::-webkit-scrollbar-thumb { background: var(--border-light); border-radius: 3px; }
  .live-transcript .interim { color: var(--text-dim); font-style: italic; }
  .live-transcript .final { color: var(--text); }
  .live-source-selector {
    display: flex; gap: 0.5rem; margin-bottom: 1.25rem;
  }
  .source-btn {
    padding: 0.4rem 0.9rem; font-size: 0.75rem; border-radius: 20px;
    border: 1px solid var(--border); background: var(--surface);
    color: var(--text-secondary); cursor: pointer; transition: all 0.2s;
  }
  .source-btn:hover { border-color: var(--accent-dim); color: var(--accent); }
  .source-btn.active { border-color: var(--accent); background: var(--accent-dim); color: var(--accent); font-weight: 600; }
  .livekit-section {
    width: 100%; margin-bottom: 1rem; background: var(--surface);
    border: 1px solid var(--border); border-radius: 12px; padding: 1rem;
  }
  .livekit-row { display: flex; gap: 0.5rem; align-items: center; }
  .livekit-input {
    flex: 1; padding: 0.4rem 0.75rem; background: var(--surface-raised);
    border: 1px solid var(--border); border-radius: 8px; color: var(--text);
    font-size: 0.85rem; outline: none;
  }
  .livekit-input:focus { border-color: var(--accent-dim); }
  .livekit-rooms { display: flex; flex-wrap: wrap; gap: 0.4rem; margin-top: 0.6rem; font-size: 0.8rem; color: var(--text-dim); }
  .room-chip {
    padding: 0.2rem 0.6rem; border-radius: 12px; border: 1px solid var(--border);
    background: var(--surface-raised); color: var(--text-secondary); cursor: pointer; font-size: 0.75rem;
  }
  .room-chip:hover { border-color: var(--accent-dim); color: var(--accent); }
  .icon-btn {
    padding: 0.4rem 0.6rem; border-radius: 8px; border: 1px solid var(--border);
    background: var(--surface-raised); color: var(--text-secondary); cursor: pointer;
    font-size: 0.9rem; line-height: 1;
  }
  .icon-btn:hover { border-color: var(--accent-dim); color: var(--accent); }
  .saved-meeting-row {
    display: flex; align-items: center; gap: 0.75rem; margin-top: 0.75rem;
    padding: 0.6rem 0.9rem; background: color-mix(in srgb, var(--green) 10%, transparent);
    border: 1px solid color-mix(in srgb, var(--green) 30%, transparent);
    border-radius: 8px; font-size: 0.82rem; color: var(--green);
  }
  .small-btn { padding: 0.25rem 0.6rem !important; font-size: 0.75rem !important; }
  .live-actions {
    display: flex; gap: 0.5rem; margin-top: 1rem; width: 100%; justify-content: center;
  }
  .live-actions .accent-btn { padding: 0.5rem 1.25rem; font-size: 0.8rem; }
  .live-actions .accent-btn:disabled { opacity: 0.35; cursor: not-allowed; }

  /* Documents tab */
  .doc-upload-zone {
    border: 1.5px dashed var(--border-light); border-radius: 12px;
    padding: 2rem 1.5rem; text-align: center; cursor: pointer;
    transition: all 0.3s; background: var(--surface); margin-bottom: 1.25rem;
  }
  .doc-upload-zone:hover, .doc-upload-zone.dragover {
    border-color: var(--accent-dim); background: var(--surface-raised);
  }
  .doc-upload-zone input { display: none; }
  .doc-upload-hint { font-size: 0.85rem; color: var(--text-dim); margin-top: 0.3rem; }
  .doc-upload-status {
    font-family: 'JetBrains Mono', monospace; font-size: 0.78rem;
    color: var(--accent); margin-top: 0.5rem;
  }
  .doc-table { width: 100%; border-collapse: collapse; }
  .doc-table th {
    font-family: 'JetBrains Mono', monospace; font-size: 0.68rem;
    text-transform: uppercase; letter-spacing: 0.07em; color: var(--text-dim);
    text-align: left; padding: 0.4rem 0.5rem; border-bottom: 1px solid var(--border);
  }
  .doc-table td {
    font-size: 0.85rem; padding: 0.6rem 0.5rem;
    border-bottom: 1px solid var(--border); vertical-align: middle;
    color: var(--text);
  }
  .doc-table tr:last-child td { border-bottom: none; }
  .doc-table tr:hover td { background: var(--surface); }
  .doc-name { font-size: 0.88rem; color: var(--text); word-break: break-word; }
  .doc-size { font-family: 'JetBrains Mono', monospace; font-size: 0.72rem; color: var(--text-dim); white-space: nowrap; }
  .doc-type-badge {
    font-family: 'JetBrains Mono', monospace; font-size: 0.65rem; text-transform: uppercase;
    padding: 0.15rem 0.5rem; border-radius: 4px; border: 1px solid var(--border);
    color: var(--text-dim); white-space: nowrap;
  }
  .doc-status-extracted { color: var(--green); font-family: 'JetBrains Mono', monospace; font-size: 0.7rem; }
  .doc-status-pending { color: var(--text-dim); font-family: 'JetBrains Mono', monospace; font-size: 0.7rem; }
  .doc-action-btn {
    background: none; border: 1px solid var(--border); color: var(--text-dim);
    font-family: 'JetBrains Mono', monospace; font-size: 0.68rem;
    padding: 0.2rem 0.55rem; border-radius: 5px; cursor: pointer;
    transition: all 0.2s; margin-right: 0.25rem;
  }
  .doc-action-btn:hover { border-color: var(--accent-dim); color: var(--accent); }
  .doc-action-btn.del:hover { border-color: rgba(196,112,112,0.5); color: var(--red); }
  .doc-context-section {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 12px; padding: 0.75rem 1rem; margin-bottom: 1rem;
  }
  .doc-context-header {
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 0.5rem;
  }
  .doc-ctx-list { display: flex; flex-direction: column; gap: 0.2rem; max-height: 140px; overflow-y: auto; }
  .doc-ctx-item {
    display: flex; align-items: center; gap: 0.5rem; padding: 0.25rem;
    border-radius: 5px; font-size: 0.82rem;
  }
  .doc-ctx-item:hover { background: var(--surface-raised); }
  .doc-ctx-item input[type="checkbox"] { accent-color: var(--accent); cursor: pointer; flex-shrink: 0; }
  .doc-ctx-name { flex: 1; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; color: var(--text); }
  .doc-ctx-ext { font-family: 'JetBrains Mono', monospace; font-size: 0.65rem; color: var(--text-dim); flex-shrink: 0; }

  footer {
    margin-top: 4rem; padding-top: 1.5rem; border-top: 1px solid var(--border);
    font-size: 0.78rem; color: var(--text-dim); text-align: center;
  }
</style>
</head>
<body>
<div class="page">

  <!-- User area (logout) -->
  <div class="user-area" id="userArea" style="display:none">
    <span class="user-name" id="userName"></span>
    <button class="logout-btn" id="logoutBtn">Sign out</button>
  </div>

  <!-- ===== Landing View ===== -->
  <div id="landingView">
    <header>
      <div class="logo">
        <div class="logo-icon">
          <svg viewBox="0 0 24 24" fill="none" stroke="#1a1714" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
            <path d="M12 2a3 3 0 0 0-3 3v7a3 3 0 0 0 6 0V5a3 3 0 0 0-3-3Z"/>
            <path d="M19 10v2a7 7 0 0 1-14 0v-2"/><line x1="12" y1="19" x2="12" y2="22"/>
          </svg>
        </div>
        <h1>Meeting Analyzer</h1>
      </div>
      <p class="tagline">Upload a recording. Get the transcript, a summary, and every action item.</p>
    </header>

    <div class="section-header">
      <span class="section-title">Workspaces</span>
      <button class="accent-btn" id="newWsBtn">+ New</button>
    </div>
    <div id="newWsForm" class="new-ws-form" style="display:none">
      <input type="text" id="newWsInput" class="new-ws-input" placeholder="Workspace name..." maxlength="80">
      <button class="new-ws-submit" id="newWsSubmit">Create</button>
      <button class="new-ws-cancel" id="newWsCancel">Cancel</button>
    </div>
    <div id="wsGrid" class="ws-grid"></div>

    <div id="unorgSection" style="display:none">
      <div class="section-header">
        <span class="section-title">Unorganized Meetings</span>
      </div>
      <div id="unorgList"></div>
    </div>
  </div>

  <!-- ===== Workspace View ===== -->
  <div id="workspaceView" style="display:none">
    <div class="ws-nav">
      <button class="back-btn" id="backBtn">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="15 18 9 12 15 6"/></svg>
        Back
      </button>
      <h2 id="wsName" class="ws-name"></h2>
      <button class="del-ws-btn" id="delWsBtn">Delete</button>
    </div>

    <div class="tabs" id="wsTabs">
      <button class="tab active" data-tab="wsUploadTab">Upload</button>
      <button class="tab" data-tab="wsLiveTab">Live</button>
      <button class="tab" data-tab="wsMeetingsTab">Meetings</button>
      <button class="tab" data-tab="wsDocumentsTab">Documents</button>
      <button class="tab" data-tab="wsChatTab">Chat</button>
    </div>

    <!-- Upload Tab -->
    <div class="tab-content active" id="wsUploadTab">
      <div id="uploadSection">
        <div class="upload-zone" id="dropZone">
          <div class="upload-icon">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round">
              <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/>
              <polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/>
            </svg>
          </div>
          <div class="upload-title">Drop your recording here</div>
          <div class="upload-hint">MP4, M4A, or MP3 &middot; or click to browse</div>
          <input type="file" id="fileInput" accept=".mp4,.m4a,.mp3,video/mp4,audio/mp4,audio/x-m4a,audio/mpeg">
        </div>
        <div class="file-chip" id="fileName"></div>
      </div>

      <div class="pipeline" id="pipeline">
        <div class="pipe-steps">
          <div class="pipe-step" data-step="upload">
            <div class="pipe-line"><div class="pipe-dot"></div><div class="pipe-connector"></div></div>
            <div class="pipe-label">Uploading file</div>
          </div>
          <div class="pipe-step" data-step="extract">
            <div class="pipe-line"><div class="pipe-dot"></div><div class="pipe-connector"></div></div>
            <div class="pipe-label">Extracting audio</div>
          </div>
          <div class="pipe-step" data-step="transcribe">
            <div class="pipe-line"><div class="pipe-dot"></div><div class="pipe-connector"></div></div>
            <div class="pipe-label">Transcribing with Whisper</div>
          </div>
          <div class="pipe-step" data-step="analyze">
            <div class="pipe-line"><div class="pipe-dot"></div></div>
            <div class="pipe-label">Analyzing with Claude</div>
          </div>
        </div>
      </div>

      <div class="error-box" id="error">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="12" r="10"/><line x1="15" y1="9" x2="9" y2="15"/><line x1="9" y1="9" x2="15" y2="15"/></svg>
        <span id="errorText"></span>
      </div>

      <div class="results" id="results">
        <div class="result-section">
          <div class="label">Summary <span class="label-line"></span></div>
          <p id="summary"></p>
          <button class="tts-btn" onclick="toggleTTS('summary', this)">Read Aloud</button>
        </div>
        <div class="result-section">
          <div class="label">Action Items <span class="label-line"></span></div>
          <ul id="actionItems"></ul>
        </div>
        <div class="result-section">
          <div class="label">Email Draft <span class="label-line"></span></div>
          <div class="email-box" id="emailBody"></div>
          <button class="tts-btn" onclick="toggleTTS('emailBody', this)">Read Aloud</button>
          <button class="copy-btn" id="copyEmailBtn">Copy email</button>
        </div>
        <div class="result-section">
          <div class="label">Full Transcript <span class="label-line"></span></div>
          <div class="transcript-box" id="transcript"></div>
        </div>
        <div class="reset-row">
          <button class="reset-btn" id="resetBtn">Analyze another recording</button>
        </div>
      </div>
    </div>

    <!-- Live Transcribe Tab -->
    <div class="tab-content" id="wsLiveTab">
      <div class="live-container">
        <!-- Source selector -->
        <div class="live-source-selector" id="liveSourceSelector">
          <button class="source-btn active" data-source="mic">🎤 Microphone</button>
          <button class="source-btn" data-source="screen">🖥️ Share Tab</button>
          <button class="source-btn" data-source="livekit">🔴 LiveKit Room</button>
        </div>
        <!-- LiveKit room picker -->
        <div class="livekit-section" id="livekitSection" style="display:none">
          <div class="livekit-row">
            <input class="livekit-input" type="text" id="livekitRoomInput" placeholder="Room name (e.g. meeting)" value="meeting" />
            <button class="icon-btn" id="livekitLoadRoomsBtn" title="Refresh room list">↻</button>
          </div>
          <div class="livekit-rooms" id="livekitRoomList"></div>
        </div>
        <button class="mic-btn" id="micBtn" title="Start recording">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <path d="M12 2a3 3 0 0 0-3 3v7a3 3 0 0 0 6 0V5a3 3 0 0 0-3-3Z"/>
            <path d="M19 10v2a7 7 0 0 1-14 0v-2"/><line x1="12" y1="19" x2="12" y2="22"/>
          </svg>
        </button>
        <div class="mic-label" id="micLabel">Click to start</div>
        <div class="live-transcript" id="liveTranscript" style="display:none"></div>
        <div class="live-actions" id="liveActions" style="display:none">
          <button class="accent-btn" id="analyzeTranscriptBtn" disabled>Analyze This</button>
          <button class="accent-btn" id="clearTranscriptBtn">Clear</button>
        </div>
        <div class="pipeline" id="livePipeline">
          <div class="pipe-steps">
            <div class="pipe-step" data-step="live-analyze">
              <div class="pipe-line"><div class="pipe-dot"></div></div>
              <div class="pipe-label">Analyzing with Claude</div>
            </div>
          </div>
        </div>
        <div class="error-box" id="liveError">
          <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="12" r="10"/><line x1="15" y1="9" x2="9" y2="15"/><line x1="9" y1="9" x2="15" y2="15"/></svg>
          <span id="liveErrorText"></span>
        </div>
        <div class="results" id="liveResults">
          <div class="result-section">
            <div class="label">Summary <span class="label-line"></span></div>
            <p id="liveSummary"></p>
          </div>
          <div class="result-section">
            <div class="label">Action Items <span class="label-line"></span></div>
            <ul id="liveActionItems"></ul>
          </div>
          <div class="result-section">
            <div class="label">Email Draft <span class="label-line"></span></div>
            <div class="email-box" id="liveEmailBody"></div>
            <button class="copy-btn" id="copyLiveEmailBtn">Copy email</button>
          </div>
          <div class="saved-meeting-row" id="savedMeetingRow" style="display:none">
            <span>&#10003; Saved as meeting</span>
            <button class="accent-btn small-btn" id="viewSavedMeetingBtn">View in Meetings &rarr;</button>
          </div>
          <div class="reset-row">
            <button class="reset-btn" id="liveResetBtn">New transcription</button>
          </div>
        </div>
      </div>
    </div>

    <!-- Meetings Tab -->
    <div class="tab-content" id="wsMeetingsTab">
      <div id="wsMeetingsList"></div>
    </div>

    <!-- Documents Tab -->
    <div class="tab-content" id="wsDocumentsTab">
      <div class="doc-upload-zone" id="docDropZone">
        <div class="upload-icon">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round">
            <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
            <polyline points="14 2 14 8 20 8"/><line x1="12" y1="18" x2="12" y2="12"/><line x1="9" y1="15" x2="15" y2="15"/>
          </svg>
        </div>
        <div class="upload-title" style="font-size:1rem">Drop a document here</div>
        <div class="doc-upload-hint">PDF, DOCX, XLSX, TXT, CSV, MD &middot; or click to browse</div>
        <div class="doc-upload-status" id="docUploadStatus"></div>
        <input type="file" id="docFileInput" accept=".pdf,.doc,.docx,.xls,.xlsx,.txt,.csv,.md,.log,.json,.yaml,.yml">
      </div>
      <div id="docListContainer">
        <div class="empty-state" id="docEmptyState">No documents uploaded yet.</div>
        <table class="doc-table" id="docTable" style="display:none">
          <thead>
            <tr>
              <th>Name</th>
              <th>Size</th>
              <th>Type</th>
              <th>Uploaded</th>
              <th>Text</th>
              <th></th>
            </tr>
          </thead>
          <tbody id="docTableBody"></tbody>
        </table>
      </div>
    </div>

    <!-- Chat Tab -->
    <div class="tab-content" id="wsChatTab">
      <div class="chat-container">
        <div class="meeting-selector" id="meetingSelector">
          <div class="ms-header">
            <span class="label">Meeting Context <span class="label-line"></span></span>
          </div>
          <div id="msList" class="ms-list"></div>
        </div>
        <div class="doc-context-section" id="docContextSection" style="display:none">
          <div class="doc-context-header">
            <span class="label">Document Context <span class="label-line"></span></span>
          </div>
          <div id="docCtxList" class="doc-ctx-list">
            <div class="empty-state" style="padding:0.3rem;font-size:0.8rem">No documents with extracted text.</div>
          </div>
        </div>
        <div id="chatMessages" class="chat-messages">
          <div class="chat-empty">Ask a question about the meetings in this workspace</div>
        </div>
        <div class="chat-input-area" id="chatInputArea">
          <textarea id="chatInput" placeholder="Ask about your meetings..." rows="1"></textarea>
          <button id="voiceToggle" class="voice-toggle" title="Voice mode: read response aloud">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <polygon points="11 5 6 9 2 9 2 15 6 15 11 19 11 5"/><path d="M19.07 4.93a10 10 0 0 1 0 14.14"/><path d="M15.54 8.46a5 5 0 0 1 0 7.07"/>
            </svg>
          </button>
          <button id="chatSendBtn" class="chat-send-btn">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
              <line x1="22" y1="2" x2="11" y2="13"/><polygon points="22 2 15 22 11 13 2 9 22 2"/>
            </svg>
          </button>
        </div>
      </div>
    </div>
  </div>

  <footer>Whisper STT &middot; Meeting Analyzer</footer>
</div>

<!-- Detail Modal -->
<div class="detail-overlay" id="detailOverlay">
  <div class="detail-panel" id="detailPanel">
    <button class="detail-close" id="detailClose">&times;</button>
    <div class="detail-title" id="detailTitle"></div>
    <div class="detail-date" id="detailDate"></div>
    <div id="detailContent"></div>
  </div>
</div>

<script>

/* --- User Auth --- */
async function loadUserInfo() {
  try {
    const resp = await fetch('/auth/user');
    if (resp.ok) {
      const user = await resp.json();
      document.getElementById('userName').textContent = user.name || user.preferred_username || user.email || 'User';
      document.getElementById('userArea').style.display = 'flex';
    }
  } catch (e) { /* Not logged in or error - hide user area */ }
}
document.getElementById('logoutBtn').addEventListener('click', function() {
  window.location.href = '/auth/logout';
});
loadUserInfo();

/* --- State --- */
let currentWorkspaceId = null;
let currentWorkspaceName = '';

/* --- Elements --- */
const dropZone = document.getElementById('dropZone');
const fileInput = document.getElementById('fileInput');
const pipeline = document.getElementById('pipeline');
const errorDiv = document.getElementById('error');
const errorText = document.getElementById('errorText');
const resultsDiv = document.getElementById('results');
const fileNameDiv = document.getElementById('fileName');
const uploadSection = document.getElementById('uploadSection');

const STEP_MAP = {
  'Saving uploaded file...': 'upload',
  'Extracting audio with ffmpeg...': 'extract',
  'Transcribing with Whisper (this may take a while)...': 'transcribe',
  'Analyzing with Claude...': 'analyze',
};

function esc(s) { const d = document.createElement('div'); d.textContent = s || ''; return d.innerHTML; }

function fmtDate(iso) {
  const d = new Date(iso);
  return d.toLocaleDateString('en-US', { month: 'short', day: 'numeric', year: 'numeric' });
}

function fmtDateTime(iso) {
  const d = new Date(iso);
  return d.toLocaleDateString('en-US', { month: 'short', day: 'numeric', year: 'numeric' })
    + ' at ' + d.toLocaleTimeString('en-US', { hour: '2-digit', minute: '2-digit' });
}

function renderMeetingCard(m) {
  return '<div class="meeting-card" data-mid="' + m.id + '">' +
    '<div class="mc-title">' + esc(m.title) + '</div>' +
    '<div class="mc-meta">' + fmtDateTime(m.date) + (m.filename ? ' &middot; ' + esc(m.filename) : '') + '</div>' +
    '<div class="mc-summary">' + esc(m.summary) + '</div>' +
  '</div>';
}

/* --- Navigation --- */
function showLanding() {
  currentWorkspaceId = null;
  currentWorkspaceName = '';
  document.getElementById('landingView').style.display = '';
  document.getElementById('workspaceView').style.display = 'none';
  loadLanding();
}

function openWorkspace(id, name) {
  currentWorkspaceId = id;
  currentWorkspaceName = name;
  document.getElementById('landingView').style.display = 'none';
  document.getElementById('workspaceView').style.display = '';
  document.getElementById('wsName').textContent = name;
  /* Reset to Upload tab */
  document.querySelectorAll('#wsTabs .tab').forEach(function(t) { t.classList.remove('active'); });
  document.querySelectorAll('#workspaceView > .tab-content').forEach(function(c) { c.classList.remove('active'); });
  document.querySelector('#wsTabs .tab[data-tab="wsUploadTab"]').classList.add('active');
  document.getElementById('wsUploadTab').classList.add('active');
  resetUpload();
  /* Clear chat */
  document.getElementById('chatMessages').innerHTML = '<div class="chat-empty">Ask a question about the meetings in this workspace</div>';
}

/* --- Landing --- */
async function loadLanding() {
  /* Load workspaces */
  const grid = document.getElementById('wsGrid');
  try {
    const resp = await fetch('/workspaces');
    const workspaces = await resp.json();
    if (workspaces.length === 0) {
      grid.innerHTML = '<div class="empty-state">No workspaces yet. Create one to organize your meetings.</div>';
    } else {
      grid.innerHTML = workspaces.map(function(w) {
        return '<div class="ws-card" data-ws-id="' + w.id + '" data-ws-name="' + esc(w.name) + '">' +
          '<div class="ws-card-name">' + esc(w.name) + '</div>' +
          '<div class="ws-card-count">' + w.meeting_count + ' meeting' + (w.meeting_count !== 1 ? 's' : '') + '</div>' +
        '</div>';
      }).join('');
    }
  } catch (e) {
    grid.innerHTML = '<div class="empty-state">Failed to load workspaces.</div>';
  }
  /* Load unorganized meetings */
  try {
    const resp = await fetch('/meetings?unorganized=true');
    const meetings = await resp.json();
    const section = document.getElementById('unorgSection');
    const list = document.getElementById('unorgList');
    if (meetings.length === 0) {
      section.style.display = 'none';
    } else {
      section.style.display = '';
      list.innerHTML = meetings.map(renderMeetingCard).join('');
    }
  } catch (e) { /* ignore */ }
}

/* Workspace grid click delegation */
document.getElementById('wsGrid').addEventListener('click', function(e) {
  const card = e.target.closest('.ws-card');
  if (card) openWorkspace(parseInt(card.dataset.wsId), card.dataset.wsName);
});

/* Unorganized meeting click delegation */
document.getElementById('unorgList').addEventListener('click', function(e) {
  const card = e.target.closest('.meeting-card');
  if (card) showMeeting(parseInt(card.dataset.mid));
});

/* New workspace form */
document.getElementById('newWsBtn').addEventListener('click', function() {
  document.getElementById('newWsForm').style.display = 'flex';
  document.getElementById('newWsInput').focus();
});
document.getElementById('newWsCancel').addEventListener('click', function() {
  document.getElementById('newWsForm').style.display = 'none';
  document.getElementById('newWsInput').value = '';
});
document.getElementById('newWsSubmit').addEventListener('click', submitNewWorkspace);
document.getElementById('newWsInput').addEventListener('keydown', function(e) {
  if (e.key === 'Enter') submitNewWorkspace();
  if (e.key === 'Escape') { document.getElementById('newWsForm').style.display = 'none'; this.value = ''; }
});

async function submitNewWorkspace() {
  const input = document.getElementById('newWsInput');
  const name = input.value.trim();
  if (!name) return;
  try {
    const resp = await fetch('/workspaces', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({name: name})
    });
    if (!resp.ok) throw new Error('Failed');
    input.value = '';
    document.getElementById('newWsForm').style.display = 'none';
    loadLanding();
  } catch (e) { /* ignore */ }
}

/* Delete workspace */
document.getElementById('delWsBtn').addEventListener('click', function() {
  if (!currentWorkspaceId) return;
  if (!confirm('Delete workspace "' + currentWorkspaceName + '"? Meetings will become unorganized.')) return;
  fetch('/workspaces/' + currentWorkspaceId, { method: 'DELETE' })
    .then(function() { showLanding(); })
    .catch(function() { /* ignore */ });
});

/* Back button */
document.getElementById('backBtn').addEventListener('click', showLanding);

/* --- Workspace Tabs --- */
document.querySelectorAll('#wsTabs .tab').forEach(function(tab) {
  tab.addEventListener('click', function() {
    document.querySelectorAll('#wsTabs .tab').forEach(function(t) { t.classList.remove('active'); });
    document.querySelectorAll('#workspaceView > .tab-content').forEach(function(c) { c.classList.remove('active'); });
    tab.classList.add('active');
    document.getElementById(tab.dataset.tab).classList.add('active');
    if (tab.dataset.tab === 'wsMeetingsTab') loadWsMeetings();
    if (tab.dataset.tab === 'wsChatTab') { loadChatMeetings(); loadDocCtx(); }
    if (tab.dataset.tab === 'wsDocumentsTab') loadDocuments();
  });
});

/* --- Workspace Meetings Tab --- */
async function loadWsMeetings() {
  if (!currentWorkspaceId) return;
  const list = document.getElementById('wsMeetingsList');
  list.innerHTML = '<div class="empty-state">Loading...</div>';
  try {
    const resp = await fetch('/workspaces/' + currentWorkspaceId + '/meetings');
    const meetings = await resp.json();
    if (meetings.length === 0) {
      list.innerHTML = '<div class="empty-state">No meetings in this workspace yet. Upload a recording in the Upload tab.</div>';
    } else {
      list.innerHTML = meetings.map(renderMeetingCard).join('');
    }
  } catch (e) {
    list.innerHTML = '<div class="empty-state">Failed to load meetings.</div>';
  }
}

document.getElementById('wsMeetingsList').addEventListener('click', function(e) {
  const card = e.target.closest('.meeting-card');
  if (card) showMeeting(parseInt(card.dataset.mid));
});

/* --- Upload --- */
dropZone.addEventListener('click', function() { fileInput.click(); });
dropZone.addEventListener('dragover', function(e) { e.preventDefault(); dropZone.classList.add('dragover'); });
dropZone.addEventListener('dragleave', function() { dropZone.classList.remove('dragover'); });
dropZone.addEventListener('drop', function(e) {
  e.preventDefault(); dropZone.classList.remove('dragover');
  if (e.dataTransfer.files.length) handleFile(e.dataTransfer.files[0]);
});
fileInput.addEventListener('change', function() { if (fileInput.files.length) handleFile(fileInput.files[0]); });
document.getElementById('resetBtn').addEventListener('click', resetUpload);

function resetUpload() {
  clearInterval(_elapsedInterval);
  resultsDiv.classList.remove('active');
  pipeline.classList.remove('active');
  uploadSection.style.display = '';
  fileNameDiv.classList.remove('active');
  errorDiv.classList.remove('active');
  document.querySelectorAll('.pipe-step').forEach(function(s) { s.classList.remove('current','done'); });
  fileInput.value = '';
}

let _elapsedInterval = null;
function advancePipeline(stepKey) {
  clearInterval(_elapsedInterval);
  const steps = document.querySelectorAll('.pipe-step');
  let found = false;
  steps.forEach(function(s) {
    var el = s.querySelector('.pipe-elapsed');
    if (el) el.remove();
    if (s.dataset.step === stepKey) { s.classList.remove('done'); s.classList.add('current'); found = true; }
    else if (!found) { s.classList.remove('current'); s.classList.add('done'); }
  });
  var currentStep = document.querySelector('.pipe-step[data-step="' + stepKey + '"]');
  if (currentStep) {
    var label = currentStep.querySelector('.pipe-label');
    var span = document.createElement('span');
    span.className = 'pipe-elapsed';
    span.textContent = '0:00';
    label.appendChild(span);
    var start = Date.now();
    _elapsedInterval = setInterval(function() {
      var sec = Math.floor((Date.now() - start) / 1000);
      var m = Math.floor(sec / 60);
      var s = sec % 60;
      span.textContent = m + ':' + (s < 10 ? '0' : '') + s;
    }, 1000);
  }
}

async function handleFile(file) {
  const ext = file.name.toLowerCase(); if (!ext.endsWith('.mp4') && !ext.endsWith('.m4a') && !ext.endsWith('.mp3')) { showError('Please upload an MP4, M4A, or MP3 file.'); return; }
  errorDiv.classList.remove('active');
  resultsDiv.classList.remove('active');
  fileNameDiv.textContent = file.name;
  fileNameDiv.classList.add('active');
  pipeline.classList.add('active');
  uploadSection.style.display = 'none';
  document.querySelectorAll('.pipe-step').forEach(function(s) { s.classList.remove('current','done'); });
  advancePipeline('upload');

  const form = new FormData();
  form.append('file', file);
  const url = currentWorkspaceId ? '/analyze?workspace_id=' + currentWorkspaceId : '/analyze';

  try {
    const resp = await fetch(url, { method: 'POST', body: form });
    if (!resp.ok) {
      const errBody = await resp.json().catch(function() { return { detail: resp.statusText }; });
      throw new Error(errBody.detail || 'Server error ' + resp.status);
    }

    const reader = resp.body.getReader();
    const decoder = new TextDecoder();
    let buffer = '';
    let finalResult = null;

    while (true) {
      const chunk = await reader.read();
      if (chunk.done) break;
      buffer += decoder.decode(chunk.value, { stream: true });
      const lines = buffer.split('\n');
      buffer = lines.pop();
      for (let i = 0; i < lines.length; i++) {
        if (!lines[i].trim()) continue;
        const msg = JSON.parse(lines[i]);
        if (msg.status) { const step = STEP_MAP[msg.status]; if (step) advancePipeline(step); }
        if (msg.result) finalResult = msg.result;
        if (msg.error) throw new Error(msg.error);
      }
    }
    if (buffer.trim()) {
      const msg = JSON.parse(buffer);
      if (msg.result) finalResult = msg.result;
      if (msg.error) throw new Error(msg.error);
    }

    if (!finalResult) throw new Error('No result received');
    clearInterval(_elapsedInterval);
    document.querySelectorAll('.pipe-step').forEach(function(s) { s.classList.remove('current'); s.classList.add('done'); });
    setTimeout(function() { pipeline.classList.remove('active'); showResults(finalResult); }, 600);
  } catch (err) {
    clearInterval(_elapsedInterval);
    showError(err.message);
    pipeline.classList.remove('active');
    uploadSection.style.display = '';
  }
}

function showError(msg) { errorText.textContent = msg; errorDiv.classList.add('active'); }

function showResults(data) {
  document.getElementById('summary').textContent = data.summary || 'No summary generated.';
  const ul = document.getElementById('actionItems');
  ul.innerHTML = '';
  (data.action_items || []).forEach(function(item) {
    const li = document.createElement('li'); li.textContent = item; ul.appendChild(li);
  });
  if (!data.action_items || data.action_items.length === 0) {
    const li = document.createElement('li'); li.textContent = 'No action items identified.'; ul.appendChild(li);
  }
  document.getElementById('emailBody').textContent = data.email_body || '';
  document.getElementById('transcript').textContent = data.transcript || '';
  resultsDiv.classList.add('active');
}

document.getElementById('copyEmailBtn').addEventListener('click', function() {
  const text = document.getElementById('emailBody').textContent;
  navigator.clipboard.writeText(text).then(function() {
    const btn = document.getElementById('copyEmailBtn');
    btn.textContent = 'Copied!';
    setTimeout(function() { btn.textContent = 'Copy email'; }, 1500);
  });
});

/* --- Meeting Detail Modal --- */
async function showMeeting(id) {
  try {
    const resp = await fetch('/meetings/' + id);
    const m = await resp.json();
    document.getElementById('detailTitle').textContent = m.title;
    document.getElementById('detailDate').textContent =
      new Date(m.date).toLocaleDateString('en-US', { weekday: 'long', month: 'long', day: 'numeric', year: 'numeric' });

    let html = '';
    html += `<div class="result-section"><div class="label">Summary <span class="label-line"></span></div><p id="detailSummary">${esc(m.summary)}</p><button class="tts-btn" onclick="toggleTTS('detailSummary', this)">Read Aloud</button><div style="clear:both"></div></div>`;
    html += '<div class="result-section"><div class="label">Action Items <span class="label-line"></span></div><ul>';
    (m.action_items || []).forEach(function(item) { html += '<li>' + esc(item) + '</li>'; });
    if (!m.action_items || m.action_items.length === 0) html += '<li>No action items identified.</li>';
    html += '</ul></div>';
    if (m.email_body) {
      html += `<div class="result-section"><div class="label">Email Draft <span class="label-line"></span></div><div class="email-box" id="detailEmail">${esc(m.email_body)}</div><button class="tts-btn" onclick="toggleTTS('detailEmail', this)">Read Aloud</button><div style="clear:both"></div></div>`;
    }
    html += '<div class="result-section"><div class="label">Full Transcript <span class="label-line"></span></div><div class="transcript-box">' + esc(m.transcript) + '</div></div>';

    document.getElementById('detailContent').innerHTML = html;
    document.getElementById('detailOverlay').classList.add('active');
  } catch (e) { /* ignore */ }
}

document.getElementById('detailClose').addEventListener('click', function() {
  stopTTS();
  document.getElementById('detailOverlay').classList.remove('active');
});
document.getElementById('detailOverlay').addEventListener('click', function(e) {
  if (e.target === document.getElementById('detailOverlay')) {
    stopTTS();
    document.getElementById('detailOverlay').classList.remove('active');
  }
});

/* --- Chat Tab --- */
async function loadChatMeetings() {
  if (!currentWorkspaceId) return;
  const list = document.getElementById('msList');
  try {
    const resp = await fetch('/workspaces/' + currentWorkspaceId + '/meetings');
    const meetings = await resp.json();
    if (meetings.length === 0) {
      list.innerHTML = '<div class="empty-state" style="padding:0.5rem">No meetings yet. Upload a recording first.</div>';
      return;
    }
    list.innerHTML = meetings.map(function(m) {
      const dateStr = fmtDate(m.date);
      return '<div class="ms-item">' +
        '<input type="checkbox" checked data-mid="' + m.id + '">' +
        '<span class="ms-item-title">' + esc(m.title) + '</span>' +
        '<span class="ms-item-date">' + dateStr + '</span>' +
        '<label class="ms-transcript-label" title="Include full transcript"><input type="checkbox" data-tid="' + m.id + '"> Full text</label>' +
      '</div>';
    }).join('');
  } catch (e) {
    list.innerHTML = '<div class="empty-state" style="padding:0.5rem">Failed to load meetings.</div>';
  }
}

document.getElementById('chatSendBtn').addEventListener('click', function() { sendChat(); });
document.getElementById('chatInput').addEventListener('keydown', function(e) {
  if (e.key === 'Enter' && !e.shiftKey) { e.preventDefault(); sendChat(); }
});

async function sendChat() {
  const input = document.getElementById('chatInput');
  const question = input.value.trim();
  if (!question) return;

  /* Gather selected meetings */
  const meetingIds = [];
  const includeTranscripts = [];
  const includeDocIds = [];
  document.querySelectorAll('#msList input[data-mid]:checked').forEach(function(cb) {
    meetingIds.push(parseInt(cb.dataset.mid));
  });
  document.querySelectorAll('#msList input[data-tid]:checked').forEach(function(cb) {
    includeTranscripts.push(parseInt(cb.dataset.tid));
  });
  document.querySelectorAll('#docCtxList input[data-docid]:checked').forEach(function(cb) {
    includeDocIds.push(parseInt(cb.dataset.docid));
  });

  if (meetingIds.length === 0) {
    showError('Select at least one meeting to chat about.');
    return;
  }

  /* Clear empty state */
  const container = document.getElementById('chatMessages');
  const emptyEl = container.querySelector('.chat-empty');
  if (emptyEl) emptyEl.remove();

  /* Add user message */
  addChatMessage('user', question);
  input.value = '';
  input.disabled = true;
  document.getElementById('chatSendBtn').disabled = true;

  /* Add assistant placeholder */
  const assistantEl = addChatMessage('assistant', '');
  assistantEl.classList.add('streaming');

  try {
    const resp = await fetch('/chat', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        meeting_ids: meetingIds,
        include_transcripts: includeTranscripts,
        include_document_ids: includeDocIds,
        question: question
      })
    });

    if (!resp.ok) {
      const errBody = await resp.json().catch(function() { return {detail: resp.statusText}; });
      throw new Error(errBody.detail || 'Error ' + resp.status);
    }

    const reader = resp.body.getReader();
    const decoder = new TextDecoder();
    let buffer = '';
    let fullText = '';
    const contentEl = assistantEl.querySelector('.msg-content');

    while (true) {
      const chunk = await reader.read();
      if (chunk.done) break;
      buffer += decoder.decode(chunk.value, { stream: true });
      const lines = buffer.split('\n');
      buffer = lines.pop();
      for (let i = 0; i < lines.length; i++) {
        if (!lines[i].trim()) continue;
        const msg = JSON.parse(lines[i]);
        if (msg.token) {
          fullText += msg.token;
          contentEl.textContent = fullText;
          scrollChat();
        }
        if (msg.warning) {
          /* Insert warning before assistant message */
          const warn = document.createElement('div');
          warn.className = 'chat-msg warning';
          warn.textContent = msg.warning;
          container.insertBefore(warn, assistantEl);
        }
        if (msg.error) throw new Error(msg.error);
      }
    }
    if (buffer.trim()) {
      const msg = JSON.parse(buffer);
      if (msg.token) { fullText += msg.token; contentEl.textContent = fullText; }
      if (msg.error) throw new Error(msg.error);
    }
    assistantEl.classList.remove('streaming');
  } catch (err) {
    assistantEl.classList.remove('streaming');
    assistantEl.querySelector('.msg-content').textContent = 'Error: ' + err.message;
  } finally {
    input.disabled = false;
    document.getElementById('chatSendBtn').disabled = false;
    input.focus();
  }
}

function addChatMessage(role, content) {
  const container = document.getElementById('chatMessages');
  const div = document.createElement('div');
  div.className = 'chat-msg ' + role;
  const contentEl = document.createElement('div');
  contentEl.className = 'msg-content';
  contentEl.textContent = content;
  div.appendChild(contentEl);
  container.appendChild(div);
  scrollChat();
  return div;
}

function scrollChat() {
  const c = document.getElementById('chatMessages');
  c.scrollTop = c.scrollHeight;
}

/* --- TTS --- */
let _ttsAudio = null;
let _ttsActiveBtn = null;

function stopTTS() {
  if (_ttsAudio) {
    _ttsAudio.pause();
    _ttsAudio.src = '';
    _ttsAudio = null;
  }
  if (_ttsActiveBtn) {
    _ttsActiveBtn.textContent = 'Read Aloud';
    _ttsActiveBtn.classList.remove('playing');
    _ttsActiveBtn = null;
  }
}

async function toggleTTS(elementId, btn) {
  /* If already playing this element, stop */
  if (_ttsActiveBtn === btn && _ttsAudio) {
    stopTTS();
    return;
  }
  /* Stop any existing playback */
  stopTTS();

  var el = document.getElementById(elementId);
  if (!el) return;
  var text = el.textContent.trim();
  if (!text) return;

  btn.textContent = 'Loading...';
  btn.classList.add('playing');
  _ttsActiveBtn = btn;

  try {
    var resp = await fetch('/tts', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({text: text})
    });
    if (!resp.ok) throw new Error('TTS failed: ' + resp.status);
    var blob = await resp.blob();
    var url = URL.createObjectURL(blob);
    _ttsAudio = new Audio(url);
    _ttsAudio.onended = function() {
      URL.revokeObjectURL(url);
      stopTTS();
    };
    _ttsAudio.onerror = function() {
      URL.revokeObjectURL(url);
      stopTTS();
    };
    btn.textContent = 'Stop';
    await _ttsAudio.play();
  } catch (e) {
    stopTTS();
    btn.textContent = 'Read Aloud';
  }
}

/* --- Streaming TTS (Voice Mode for Chat) --- */
let _voiceMode = false;
let _ttsStreamWs = null;
let _ttsAudioCtx = null;
let _ttsPlayerNode = null;

const _workletCode = `
class PCMPlayerProcessor extends AudioWorkletProcessor {
  constructor() {
    super();
    this.bufSize = sampleRate * 60;
    this.buf = new Float32Array(this.bufSize);
    this.wr = 0;
    this.rd = 0;
    this.port.onmessage = (e) => {
      if (e.data && e.data.type === 'clear') { this.rd = this.wr; return; }
      const s = new Int16Array(e.data);
      for (let i = 0; i < s.length; i++) {
        this.buf[this.wr] = s[i] / 32768.0;
        this.wr = (this.wr + 1) % this.bufSize;
        if (this.wr === this.rd) this.rd = (this.rd + 1) % this.bufSize;
      }
    };
  }
  process(inputs, outputs) {
    const ch = outputs[0][0];
    for (let i = 0; i < ch.length; i++) {
      if (this.rd !== this.wr) {
        ch[i] = this.buf[this.rd];
        this.rd = (this.rd + 1) % this.bufSize;
      } else {
        ch[i] = 0;
      }
    }
    return true;
  }
}
registerProcessor('pcm-player-processor', PCMPlayerProcessor);
`;

async function initStreamingAudio(sampleRate) {
  if (_ttsAudioCtx) {
    _ttsPlayerNode.port.postMessage({ type: 'clear' });
    if (_ttsAudioCtx.sampleRate === sampleRate) return;
    _ttsAudioCtx.close();
  }
  _ttsAudioCtx = new AudioContext({ sampleRate: sampleRate });
  const blob = new Blob([_workletCode], { type: 'application/javascript' });
  const url = URL.createObjectURL(blob);
  await _ttsAudioCtx.audioWorklet.addModule(url);
  URL.revokeObjectURL(url);
  _ttsPlayerNode = new AudioWorkletNode(_ttsAudioCtx, 'pcm-player-processor');
  _ttsPlayerNode.connect(_ttsAudioCtx.destination);
}

function stopStreamingAudio() {
  if (_ttsStreamWs) {
    try { _ttsStreamWs.close(); } catch(e) {}
    _ttsStreamWs = null;
  }
  if (_ttsPlayerNode) {
    _ttsPlayerNode.port.postMessage({ type: 'clear' });
  }
}

document.getElementById('voiceToggle').addEventListener('click', function() {
  _voiceMode = !_voiceMode;
  this.classList.toggle('active', _voiceMode);
  this.title = _voiceMode ? 'Voice mode ON' : 'Voice mode: read response aloud';
  if (!_voiceMode) stopStreamingAudio();
});

/* Override sendChat when voice mode is on */
const _origSendChat = sendChat;

async function sendChatVoice() {
  const input = document.getElementById('chatInput');
  const question = input.value.trim();
  if (!question) return;

  const meetingIds = [];
  const includeTranscripts = [];
  const includeDocIds = [];
  document.querySelectorAll('#msList input[data-mid]:checked').forEach(function(cb) {
    meetingIds.push(parseInt(cb.dataset.mid));
  });
  document.querySelectorAll('#msList input[data-tid]:checked').forEach(function(cb) {
    includeTranscripts.push(parseInt(cb.dataset.tid));
  });
  document.querySelectorAll('#docCtxList input[data-docid]:checked').forEach(function(cb) {
    includeDocIds.push(parseInt(cb.dataset.docid));
  });

  if (meetingIds.length === 0) {
    showError('Select at least one meeting to chat about.');
    return;
  }

  const container = document.getElementById('chatMessages');
  const emptyEl = container.querySelector('.chat-empty');
  if (emptyEl) emptyEl.remove();

  addChatMessage('user', question);
  input.value = '';
  input.disabled = true;
  document.getElementById('chatSendBtn').disabled = true;

  const assistantEl = addChatMessage('assistant', '');
  assistantEl.classList.add('streaming');
  const contentEl = assistantEl.querySelector('.msg-content');
  let fullText = '';

  stopStreamingAudio();

  const proto = location.protocol === 'https:' ? 'wss:' : 'ws:';
  const ws = new WebSocket(proto + '//' + location.host + '/ws/tts-stream');
  _ttsStreamWs = ws;

  ws.onopen = function() {
    ws.send(JSON.stringify({
      action: 'start_chat',
      question: question,
      meeting_ids: meetingIds,
      include_transcripts: includeTranscripts,
      include_document_ids: includeDocIds,
    }));
  };

  ws.onmessage = async function(event) {
    if (typeof event.data === 'string') {
      const msg = JSON.parse(event.data);
      switch (msg.type) {
        case 'audio_config':
          await initStreamingAudio(msg.sample_rate);
          await _ttsAudioCtx.resume();
          break;
        case 'token':
          fullText += msg.text;
          contentEl.textContent = fullText;
          scrollChat();
          break;
        case 'sentence':
          break;
        case 'warning':
          var warn = document.createElement('div');
          warn.className = 'chat-msg warning';
          warn.textContent = msg.message;
          container.insertBefore(warn, assistantEl);
          break;
        case 'done':
          assistantEl.classList.remove('streaming');
          break;
        case 'error':
          assistantEl.classList.remove('streaming');
          contentEl.textContent = 'Error: ' + msg.message;
          break;
      }
    } else {
      /* Binary: raw PCM audio */
      if (_ttsPlayerNode) {
        _ttsPlayerNode.port.postMessage(event.data);
      }
    }
  };

  ws.onerror = function() {
    assistantEl.classList.remove('streaming');
    if (!fullText) contentEl.textContent = 'Error: WebSocket connection failed';
  };

  ws.onclose = function() {
    _ttsStreamWs = null;
    assistantEl.classList.remove('streaming');
    input.disabled = false;
    document.getElementById('chatSendBtn').disabled = false;
    input.focus();
  };
}

/* Patch sendChat to route through voice mode when active */
sendChat = function() {
  if (_voiceMode) return sendChatVoice();
  return _origSendChat();
};

/* --- Live Transcribe --- */
let liveWs = null;
let liveAudioCtx = null;
let liveStream = null;
let liveWorklet = null;
let liveRecording = false;
let liveSegments = {}; /* keyed by segment index for dedup */
let liveFinalText = '';
let liveSource = 'mic'; /* 'mic' | 'screen' | 'livekit' */
let livekitRoom = null; /* LiveKit Room instance */

/* Load a script dynamically (idempotent) */
function loadScript(src) {
  return new Promise(function(res, rej) {
    if (document.querySelector('script[src="' + src + '"]')) { res(); return; }
    const s = document.createElement('script');
    s.src = src; s.onload = res; s.onerror = rej;
    document.head.appendChild(s);
  });
}

/* Source selector */
document.querySelectorAll('#liveSourceSelector .source-btn').forEach(function(btn) {
  btn.addEventListener('click', function() {
    if (liveRecording) return;
    document.querySelectorAll('#liveSourceSelector .source-btn').forEach(function(b) { b.classList.remove('active'); });
    btn.classList.add('active');
    liveSource = btn.dataset.source;
    document.getElementById('livekitSection').style.display = liveSource === 'livekit' ? '' : 'none';
  });
});

/* LiveKit room list refresh */
document.getElementById('livekitLoadRoomsBtn').addEventListener('click', async function() {
  const listEl = document.getElementById('livekitRoomList');
  listEl.textContent = 'Loading...';
  try {
    const resp = await fetch('/livekit/rooms');
    const data = await resp.json();
    if (!data.rooms || data.rooms.length === 0) {
      listEl.textContent = 'No active rooms';
      return;
    }
    listEl.innerHTML = '';
    data.rooms.forEach(function(name) {
      const btn = document.createElement('button');
      btn.className = 'room-chip';
      btn.textContent = name;
      btn.addEventListener('click', function() {
        document.getElementById('livekitRoomInput').value = name;
      });
      listEl.appendChild(btn);
    });
  } catch(e) {
    listEl.textContent = 'Failed to load rooms';
  }
});

const micBtn = document.getElementById('micBtn');
const micLabel = document.getElementById('micLabel');
const liveTranscript = document.getElementById('liveTranscript');
const liveActions = document.getElementById('liveActions');
const analyzeTranscriptBtn = document.getElementById('analyzeTranscriptBtn');

micBtn.addEventListener('click', function() {
  if (liveRecording) stopLiveTranscribe(); else startLiveTranscribe();
});

async function startLiveTranscribe() {
  /* Reset UI */
  liveSegments = {};
  liveFinalText = '';
  liveTranscript.innerHTML = '';
  liveTranscript.style.display = '';
  liveActions.style.display = 'flex';
  analyzeTranscriptBtn.disabled = true;
  document.getElementById('liveResults').classList.remove('active');
  document.getElementById('livePipeline').classList.remove('active');
  document.getElementById('liveError').classList.remove('active');
  document.getElementById('savedMeetingRow').style.display = 'none';

  if (liveSource === 'livekit') {
    await startLiveKitCapture();
    return;
  }

  try {
    if (liveSource === 'screen') {
      liveStream = await navigator.mediaDevices.getDisplayMedia({ audio: true, video: true });
      if (!liveStream.getAudioTracks().length) {
        micLabel.textContent = 'No audio — check "Share tab audio"';
        return;
      }
      /* Stop capture when user clicks browser stop-sharing button */
      liveStream.getAudioTracks()[0].addEventListener('ended', function() {
        if (liveRecording) stopLiveTranscribe();
      });
    } else {
      liveStream = await navigator.mediaDevices.getUserMedia({
        audio: { sampleRate: 16000, channelCount: 1, echoCancellation: true, noiseSuppression: true }
      });
    }
  } catch (e) {
    micLabel.textContent = liveSource === 'screen' ? 'Screen share denied' : 'Mic access denied';
    return;
  }

  liveRecording = true;
  micBtn.classList.add('recording');
  micLabel.textContent = liveSource === 'screen' ? 'Capturing tab audio...' : 'Recording...';

  /* Open WebSocket to our backend */
  const proto = location.protocol === 'https:' ? 'wss:' : 'ws:';
  liveWs = new WebSocket(proto + '//' + location.host + '/ws/transcribe');
  liveWs.binaryType = 'arraybuffer';

  liveWs.onopen = function() {
    liveWs.send(JSON.stringify({ action: 'start', language: 'en' }));
  };

  liveWs.onmessage = _handleLiveWsMessage;

  liveWs.onclose = function() {
    if (liveRecording) finalizeLiveTranscript();
  };

  liveWs.onerror = function() {
    showLiveError('WebSocket connection failed. Is the transcription service running?');
    stopLiveTranscribe();
  };

  /* Set up audio capture with AudioWorklet or ScriptProcessor fallback */
  liveAudioCtx = new (window.AudioContext || window.webkitAudioContext)({ sampleRate: 16000 });
  const source = liveAudioCtx.createMediaStreamSource(liveStream);

  /* ScriptProcessor approach (widely supported) */
  const processor = liveAudioCtx.createScriptProcessor(4096, 1, 1);
  processor.onaudioprocess = function(e) {
    if (!liveRecording || !liveWs || liveWs.readyState !== WebSocket.OPEN) return;
    const input = e.inputBuffer.getChannelData(0);
    /* Send float32 PCM directly */
    const buffer = new Float32Array(input.length);
    buffer.set(input);
    liveWs.send(buffer.buffer);
  };
  source.connect(processor);
  processor.connect(liveAudioCtx.destination);
  liveWorklet = { source: source, processor: processor };
}

function renderLiveTranscript() {
  const sorted = Object.values(liveSegments).sort(function(a, b) { return a.start - b.start; });
  let html = '';
  sorted.forEach(function(seg) {
    if (seg.completed) {
      html += '<span class="final">' + esc(seg.text) + '</span> ';
    } else {
      html += '<span class="interim">' + esc(seg.text) + '</span> ';
    }
  });
  liveTranscript.innerHTML = html;
  liveTranscript.scrollTop = liveTranscript.scrollHeight;
}

function finalizeLiveTranscript() {
  liveRecording = false;
  micBtn.classList.remove('recording');
  micLabel.textContent = 'Click to start';

  /* Build final text from completed segments */
  const sorted = Object.values(liveSegments).sort(function(a, b) { return a.start - b.start; });
  liveFinalText = sorted.map(function(s) { return s.text; }).join(' ').trim();
  analyzeTranscriptBtn.disabled = !liveFinalText;
}

function stopLiveTranscribe() {
  if (liveWs && liveWs.readyState === WebSocket.OPEN) {
    liveWs.send(JSON.stringify({ action: 'stop' }));
  }

  if (liveWorklet) {
    liveWorklet.processor.disconnect();
    liveWorklet.source.disconnect();
    liveWorklet = null;
  }
  if (liveAudioCtx) {
    liveAudioCtx.close().catch(function(){});
    liveAudioCtx = null;
  }
  if (liveStream) {
    liveStream.getTracks().forEach(function(t) { t.stop(); });
    liveStream = null;
  }
  if (livekitRoom) {
    livekitRoom.disconnect();
    livekitRoom = null;
  }

  finalizeLiveTranscript();
}

/* --- LiveKit room capture --- */
async function startLiveKitCapture() {
  const roomName = document.getElementById('livekitRoomInput').value.trim() || 'meeting';
  micLabel.textContent = 'Connecting to LiveKit...';

  /* Get access token from backend */
  let tokenData;
  try {
    const r = await fetch('/livekit/token', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ room: roomName }),
    });
    if (!r.ok) throw new Error('Token request failed: ' + r.status);
    tokenData = await r.json();
  } catch (e) {
    showLiveError('Failed to get LiveKit token: ' + e.message);
    micLabel.textContent = 'Click to start';
    return;
  }

  /* Load LiveKit client SDK if not already loaded (with CDN fallbacks) */
  const LIVEKIT_CDN_URLS = [
    'https://cdn.jsdelivr.net/npm/livekit-client@2/dist/livekit-client.umd.min.js',
    'https://unpkg.com/livekit-client@2/dist/livekit-client.umd.min.js',
  ];
  
  if (!window.LivekitClient) {
    micLabel.textContent = 'Loading LiveKit SDK...';
    let loaded = false;
    for (const cdnUrl of LIVEKIT_CDN_URLS) {
      try {
        await loadScript(cdnUrl);
        if (window.LivekitClient) { loaded = true; break; }
      } catch (e) { console.warn('Failed to load LiveKit SDK from', cdnUrl, e); }
    }
    if (!loaded) {
      showLiveError('Failed to load LiveKit SDK. Network may be restricted or SDK unavailable. Try using Microphone or Screen Capture instead.');
      micLabel.textContent = 'Click to start';
      return;
    }
  }

  /* Open WhisperLive WebSocket first */
  const proto = location.protocol === 'https:' ? 'wss:' : 'ws:';
  liveWs = new WebSocket(proto + '//' + location.host + '/ws/transcribe');
  liveWs.binaryType = 'arraybuffer';
  liveWs.onmessage = _handleLiveWsMessage;
  liveWs.onerror = function() { showLiveError('WebSocket connection failed.'); stopLiveTranscribe(); };
  liveWs.onclose = function() { if (liveRecording) finalizeLiveTranscript(); };

  await new Promise(function(res, rej) {
    liveWs.onopen = function() {
      liveWs.send(JSON.stringify({ action: 'start', language: 'en' }));
      res();
    };
    setTimeout(rej, 5000);
  }).catch(function() {
    showLiveError('WebSocket timed out connecting.');
    return;
  });

  /* Connect to LiveKit room as subscriber */
  const room = new LivekitClient.Room({ adaptiveStream: false, dynacast: false });

  room.on(LivekitClient.RoomEvent.TrackSubscribed, function(track) {
    if (track.kind !== LivekitClient.Track.Kind.Audio) return;
    /* Route LiveKit audio through ScriptProcessor → WhisperLive */
    const mediaStream = new MediaStream([track.mediaStreamTrack]);
    liveStream = mediaStream;
    liveAudioCtx = new (window.AudioContext || window.webkitAudioContext)({ sampleRate: 16000 });
    const source = liveAudioCtx.createMediaStreamSource(mediaStream);
    const processor = liveAudioCtx.createScriptProcessor(4096, 1, 1);
    processor.onaudioprocess = function(e) {
      if (!liveRecording || !liveWs || liveWs.readyState !== WebSocket.OPEN) return;
      const buf = new Float32Array(e.inputBuffer.getChannelData(0));
      liveWs.send(buf.buffer);
    };
    source.connect(processor);
    processor.connect(liveAudioCtx.destination);
    liveWorklet = { source: source, processor: processor };
    const n = room.remoteParticipants.size;
    micLabel.textContent = 'Listening (' + n + ' participant' + (n !== 1 ? 's' : '') + ')...';
  });

  room.on(LivekitClient.RoomEvent.ParticipantConnected, function() {
    const n = room.remoteParticipants.size;
    micLabel.textContent = 'Listening (' + n + ' participant' + (n !== 1 ? 's' : '') + ')...';
  });

  room.on(LivekitClient.RoomEvent.ParticipantDisconnected, function() {
    const n = room.remoteParticipants.size;
    micLabel.textContent = n > 0
      ? 'Listening (' + n + ' participant' + (n !== 1 ? 's' : '') + ')...'
      : 'Room empty — waiting...';
  });

  room.on(LivekitClient.RoomEvent.Disconnected, function() {
    if (liveRecording) stopLiveTranscribe();
  });

  try {
    await room.connect(tokenData.url, tokenData.token, { autoSubscribe: true });
  } catch (e) {
    showLiveError('Failed to connect to LiveKit room: ' + e.message);
    liveWs.close();
    liveWs = null;
    micLabel.textContent = 'Click to start';
    return;
  }

  livekitRoom = room;
  liveRecording = true;
  micBtn.classList.add('recording');
  const n = room.remoteParticipants.size;
  micLabel.textContent = n > 0
    ? 'Listening (' + n + ' participant' + (n !== 1 ? 's' : '') + ')...'
    : 'Joined "' + roomName + '" — waiting for audio...';
}

function _handleLiveWsMessage(evt) {
  try {
    const data = JSON.parse(evt.data);
    if (data.message === 'SERVER_READY') { micLabel.textContent = liveSource === 'livekit' ? 'Room connected...' : 'Listening...'; return; }
    if (data.status === 'WAIT') { micLabel.textContent = 'Server busy, waiting...'; return; }
    if (data.segments) {
      data.segments.forEach(function(seg, i) { liveSegments[seg.start + '-' + i] = seg; });
      renderLiveTranscript();
    }
    if (data.done) finalizeLiveTranscript();
    if (data.error) { showLiveError(data.error); stopLiveTranscribe(); }
  } catch (e) { /* ignore */ }
}

function showLiveError(msg) {
  document.getElementById('liveErrorText').textContent = msg;
  document.getElementById('liveError').classList.add('active');
}

analyzeTranscriptBtn.addEventListener('click', async function() {
  if (!liveFinalText) return;
  analyzeTranscriptBtn.disabled = true;
  analyzeTranscriptBtn.textContent = 'Analyzing...';

  document.getElementById('livePipeline').classList.add('active');
  var step = document.querySelector('#livePipeline .pipe-step');
  if (step) step.classList.add('current');

  try {
    const resp = await fetch('/analyze-text', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ text: liveFinalText, workspace_id: currentWorkspaceId })
    });
    if (!resp.ok) {
      const errBody = await resp.json().catch(function() { return { detail: resp.statusText }; });
      throw new Error(errBody.detail || 'Error ' + resp.status);
    }

    const reader = resp.body.getReader();
    const decoder = new TextDecoder();
    let buffer = '';
    let finalResult = null;

    while (true) {
      const chunk = await reader.read();
      if (chunk.done) break;
      buffer += decoder.decode(chunk.value, { stream: true });
      const lines = buffer.split('\n');
      buffer = lines.pop();
      for (let i = 0; i < lines.length; i++) {
        if (!lines[i].trim()) continue;
        const msg = JSON.parse(lines[i]);
        if (msg.result) finalResult = msg.result;
        if (msg.error) throw new Error(msg.error);
      }
    }
    if (buffer.trim()) {
      const msg = JSON.parse(buffer);
      if (msg.result) finalResult = msg.result;
      if (msg.error) throw new Error(msg.error);
    }

    if (!finalResult) throw new Error('No result received');

    document.getElementById('livePipeline').classList.remove('active');
    if (step) { step.classList.remove('current'); step.classList.add('done'); }

    document.getElementById('liveSummary').textContent = finalResult.summary || '';
    const ul = document.getElementById('liveActionItems');
    ul.innerHTML = '';
    (finalResult.action_items || []).forEach(function(item) {
      const li = document.createElement('li'); li.textContent = item; ul.appendChild(li);
    });
    if (!finalResult.action_items || finalResult.action_items.length === 0) {
      const li = document.createElement('li'); li.textContent = 'No action items identified.'; ul.appendChild(li);
    }
    document.getElementById('liveEmailBody').textContent = finalResult.email_body || '';
    document.getElementById('liveResults').classList.add('active');

    /* Show "View in Meetings" link — /analyze-text already saved the meeting */
    if (finalResult.id) {
      const savedRow = document.getElementById('savedMeetingRow');
      savedRow.style.display = '';
      document.getElementById('viewSavedMeetingBtn').onclick = function() {
        document.querySelector('[data-tab="wsMeetingsTab"]').click();
      };
    }

  } catch (err) {
    document.getElementById('livePipeline').classList.remove('active');
    showLiveError(err.message);
  } finally {
    analyzeTranscriptBtn.textContent = 'Analyze This';
    analyzeTranscriptBtn.disabled = false;
  }
});

document.getElementById('clearTranscriptBtn').addEventListener('click', function() {
  liveSegments = {};
  liveFinalText = '';
  liveTranscript.innerHTML = '';
  analyzeTranscriptBtn.disabled = true;
  document.getElementById('liveResults').classList.remove('active');
  document.getElementById('liveError').classList.remove('active');
  document.getElementById('livePipeline').classList.remove('active');
  document.getElementById('savedMeetingRow').style.display = 'none';
});

document.getElementById('copyLiveEmailBtn').addEventListener('click', function() {
  const text = document.getElementById('liveEmailBody').textContent;
  navigator.clipboard.writeText(text).then(function() {
    const btn = document.getElementById('copyLiveEmailBtn');
    btn.textContent = 'Copied!';
    setTimeout(function() { btn.textContent = 'Copy email'; }, 1500);
  });
});

document.getElementById('liveResetBtn').addEventListener('click', function() {
  liveSegments = {};
  liveFinalText = '';
  liveTranscript.innerHTML = '';
  liveTranscript.style.display = 'none';
  liveActions.style.display = 'none';
  analyzeTranscriptBtn.disabled = true;
  document.getElementById('liveResults').classList.remove('active');
  document.getElementById('liveError').classList.remove('active');
  document.getElementById('livePipeline').classList.remove('active');
});

/* --- Documents Tab --- */
const docDropZone = document.getElementById('docDropZone');
const docFileInput = document.getElementById('docFileInput');

docDropZone.addEventListener('click', function() { docFileInput.click(); });
docDropZone.addEventListener('dragover', function(e) { e.preventDefault(); docDropZone.classList.add('dragover'); });
docDropZone.addEventListener('dragleave', function() { docDropZone.classList.remove('dragover'); });
docDropZone.addEventListener('drop', function(e) {
  e.preventDefault(); docDropZone.classList.remove('dragover');
  if (e.dataTransfer.files.length) handleDocUpload(e.dataTransfer.files[0]);
});
docFileInput.addEventListener('change', function() { if (docFileInput.files.length) handleDocUpload(docFileInput.files[0]); });

function fmtBytes(bytes) {
  if (bytes < 1024) return bytes + ' B';
  if (bytes < 1024 * 1024) return (bytes / 1024).toFixed(1) + ' KB';
  return (bytes / (1024 * 1024)).toFixed(1) + ' MB';
}

function docExtension(filename) {
  const m = filename.match(/\\.([^.]+)$/);
  return m ? m[1].toUpperCase() : '?';
}

async function handleDocUpload(file) {
  if (!currentWorkspaceId) return;
  const status = document.getElementById('docUploadStatus');
  status.textContent = 'Uploading ' + file.name + '...';
  const form = new FormData();
  form.append('file', file);
  try {
    const resp = await fetch('/workspaces/' + currentWorkspaceId + '/documents', {
      method: 'POST', body: form
    });
    if (!resp.ok) {
      const err = await resp.json().catch(function() { return {detail: resp.statusText}; });
      status.textContent = 'Error: ' + (err.detail || 'Upload failed');
      return;
    }
    status.textContent = 'Uploaded. Extracting text...';
    docFileInput.value = '';
    await loadDocuments();
    setTimeout(function() { status.textContent = ''; }, 3000);
  } catch (e) {
    status.textContent = 'Upload error: ' + e.message;
  }
}

async function loadDocuments() {
  if (!currentWorkspaceId) return;
  const emptyState = document.getElementById('docEmptyState');
  const table = document.getElementById('docTable');
  const tbody = document.getElementById('docTableBody');
  try {
    const resp = await fetch('/workspaces/' + currentWorkspaceId + '/documents');
    const docs = await resp.json();
    if (docs.length === 0) {
      emptyState.style.display = '';
      table.style.display = 'none';
    } else {
      emptyState.style.display = 'none';
      table.style.display = '';
      tbody.innerHTML = docs.map(function(d) {
        const textStatus = d.has_text
          ? '<span class="doc-status-extracted">extracted</span>'
          : '<span class="doc-status-pending">pending</span>';
        return '<tr>' +
          '<td class="doc-name">' + esc(d.filename) + '</td>' +
          '<td class="doc-size">' + fmtBytes(d.file_size) + '</td>' +
          '<td><span class="doc-type-badge">' + docExtension(d.filename) + '</span></td>' +
          '<td class="doc-size">' + fmtDate(d.uploaded_at) + '</td>' +
          '<td>' + textStatus + '</td>' +
          '<td style="white-space:nowrap">' +
            '<button class="doc-action-btn" data-doc-action="download" data-doc-id="' + d.id + '">DL</button>' +
            '<button class="doc-action-btn del" data-doc-action="delete" data-doc-id="' + d.id + '">Del</button>' +
          '</td>' +
        '</tr>';
      }).join('');
    }
  } catch (e) {
    emptyState.textContent = 'Failed to load documents.';
    emptyState.style.display = '';
    table.style.display = 'none';
  }
}

async function downloadDoc(docId) {
  if (!currentWorkspaceId) return;
  const a = document.createElement('a');
  a.href = '/workspaces/' + currentWorkspaceId + '/documents/' + docId + '/download';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
}

async function deleteDoc(docId) {
  if (!currentWorkspaceId) return;
  if (!confirm('Delete this document?')) return;
  try {
    const resp = await fetch('/workspaces/' + currentWorkspaceId + '/documents/' + docId, { method: 'DELETE' });
    if (!resp.ok) return;
    await loadDocuments();
  } catch (e) { /* ignore */ }
}

document.getElementById('docTableBody').addEventListener('click', function(e) {
  const button = e.target.closest('[data-doc-action]');
  if (!button) return;
  const docId = parseInt(button.dataset.docId, 10);
  if (!docId) return;
  if (button.dataset.docAction === 'download') {
    downloadDoc(docId);
  } else if (button.dataset.docAction === 'delete') {
    deleteDoc(docId);
  }
});

/* --- Document Context in Chat --- */
async function loadDocCtx() {
  if (!currentWorkspaceId) return;
  const section = document.getElementById('docContextSection');
  const list = document.getElementById('docCtxList');
  try {
    const resp = await fetch('/workspaces/' + currentWorkspaceId + '/documents');
    const docs = await resp.json();
    const withText = docs.filter(function(d) { return d.has_text; });
    if (withText.length === 0) {
      section.style.display = 'none';
      return;
    }
    section.style.display = '';
    list.innerHTML = withText.map(function(d) {
      return '<div class="doc-ctx-item">' +
        '<input type="checkbox" data-docid="' + d.id + '">' +
        '<span class="doc-ctx-name">' + esc(d.filename) + '</span>' +
        '<span class="doc-ctx-ext">' + docExtension(d.filename) + '</span>' +
      '</div>';
    }).join('');
  } catch (e) {
    section.style.display = 'none';
  }
}

/* --- Init --- */
loadLanding();
</script>
</body>
</html>
"""


def _inject_team_chat(html: str) -> str:
    """Inject the floating team chat widget if configured."""
    if not TEAM_CHAT_HOMESERVER or not TEAM_CHAT_TOKEN or not TEAM_CHAT_ROOM_ID:
        return html

    widget_html = f"""
<style>
.team-chat-fab {{
  position: fixed;
  bottom: 24px;
  right: 24px;
  z-index: 10000;
  width: 56px;
  height: 56px;
  border-radius: 50%;
  background: #0066cc;
  color: white;
  border: none;
  cursor: pointer;
  font-size: 24px;
  box-shadow: 0 4px 12px rgba(0,0,0,0.25);
  display: flex;
  align-items: center;
  justify-content: center;
  transition: transform 0.2s;
}}
.team-chat-fab:hover {{ transform: scale(1.1); }}
.team-chat-panel {{
  position: fixed;
  bottom: 92px;
  right: 24px;
  width: 380px;
  height: 500px;
  z-index: 10000;
  display: none;
  border-radius: 12px;
  box-shadow: 0 8px 32px rgba(0,0,0,0.2);
  overflow: hidden;
}}
.team-chat-panel.open {{ display: block; }}
</style>
<button class="team-chat-fab" onclick="document.querySelector('.team-chat-panel').classList.toggle('open')" title="Team Agent Chat">&#x1f4ac;</button>
<div class="team-chat-panel"
     data-team-chat
     data-homeserver="{TEAM_CHAT_HOMESERVER}"
     data-token="{TEAM_CHAT_TOKEN}"
     data-user-id="{TEAM_CHAT_USER_ID}"
     data-room-id="{TEAM_CHAT_ROOM_ID}">
</div>
<script src="team-chat.js"></script>
"""
    return html.replace("</body>", widget_html + "\n</body>", 1)


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    # Check if user is authenticated
    user = request.session.get("user")
    if not user:
        # Redirect to login if not authenticated
        return RedirectResponse(url=f"{APP_PATH_PREFIX}/auth/login")

    prefix_script = f"<script>const __BASE='{APP_PATH_PREFIX}';const __oF=window.fetch;window.fetch=function(u,o){{if(typeof u==='string'&&u.startsWith('/')&&!u.startsWith(__BASE))u=__BASE+u;return __oF.call(this,u,o)}};const __WS=window.WebSocket;window.WebSocket=function(u,p){{return new __WS(u.replace(/^(wss?:\\/\\/[^\\/]+)(\\/ws\\/)/,'$1'+__BASE+'$2'),p)}};Object.setPrototypeOf(window.WebSocket,__WS);</script>"
    html = HTML_PAGE.replace("</head>", prefix_script + "\n</head>", 1)
    html = _inject_team_chat(html)
    return HTMLResponse(
        html,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.get("/team-chat.js")
async def team_chat_js():
    """Serve the team chat widget bundle."""
    widget_path = os.path.join(os.path.dirname(__file__), "team-chat.js")
    if not os.path.exists(widget_path):
        raise HTTPException(status_code=404, detail="Widget not found")
    with open(widget_path, "r") as f:
        return StreamingResponse(
            io.BytesIO(f.read().encode()),
            media_type="application/javascript",
            headers={"Cache-Control": "public, max-age=3600"},
        )


# ---- Authentication dependency ----


async def get_current_user(request: Request):
    """Dependency to get the current user from session. Raises 401 if not authenticated."""
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user


# ---- OAuth2/OIDC Authentication endpoints ----


@app.get("/auth/login")
async def auth_login(request: Request):
    """Redirect to Keycloak for authentication."""
    redirect_uri = KEYCLOAK_CALLBACK_URL
    return await oauth.keycloak.authorize_redirect(request, redirect_uri)


@app.get("/auth/callback")
async def auth_callback(request: Request):
    """Handle OAuth callback from Keycloak."""
    try:
        token = await oauth.keycloak.authorize_access_token(request)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Authentication failed: {str(e)}")
    
    # Get user info from the token
    user = token.get("userinfo")
    if not user:
        raise HTTPException(status_code=400, detail="Failed to get user info")
    
    # Store user info in session
    request.session["user"] = dict(user)
    request.session["access_token"] = token.get("access_token")
    request.session["refresh_token"] = token.get("refresh_token")
    
    # Redirect to the main app
    return RedirectResponse(url=f"{APP_PATH_PREFIX}/")


@app.get("/auth/logout")
async def auth_logout(request: Request):
    """Log out user and redirect to Keycloak logout."""
    # Clear the session
    request.session.clear()
    
    # Redirect to Keycloak logout endpoint
    keycloak_logout_url = (
        f"{KEYCLOAK_URL}/realms/{KEYCLOAK_REALM}/protocol/openid-connect/logout"
        f"?redirect_uri={APP_PATH_PREFIX}/"
    )
    return RedirectResponse(url=keycloak_logout_url)


@app.get("/auth/user")
async def auth_user(request: Request):
    """Return current user info from session."""
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user


# ---- TTS proxy endpoint ----


@app.post("/tts")
async def tts_proxy(body: TTSRequest):
    text = body.text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="Text is required.")

    # Use in-process Piper when available (no network hop)
    if piper_voice:
        import io
        import wave

        def _synthesize_wav():
            import numpy as np
            pcm_chunks = []
            for audio_chunk in piper_voice.synthesize(text):
                pcm_chunks.append((audio_chunk.audio_float_array * 32767).astype(np.int16).tobytes())
            pcm = b"".join(pcm_chunks)
            buf = io.BytesIO()
            with wave.open(buf, "wb") as wf:
                wf.setnchannels(1)
                wf.setsampwidth(2)
                wf.setframerate(piper_voice.config.sample_rate)
                wf.writeframes(pcm)
            return buf.getvalue()

        wav_bytes = await asyncio.to_thread(_synthesize_wav)
        return StreamingResponse(
            iter([wav_bytes]),
            media_type="audio/wav",
            headers={"Content-Disposition": "inline"},
        )

    # Fallback: proxy to piper-tts pod
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            PIPER_TTS_URL,
            json={"text": text},
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail="TTS service error.")
        return StreamingResponse(
            iter([resp.content]),
            media_type="audio/wav",
            headers={"Content-Disposition": "inline"},
        )


# ---- Workspace endpoints ----


@app.post("/workspaces")
async def create_workspace(body: WorkspaceCreate):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "INSERT INTO workspaces (name) VALUES ($1) RETURNING id, name, created_at",
            body.name,
        )
        return dict(row)


@app.get("/workspaces")
async def list_workspaces():
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT w.id, w.name, w.created_at, COUNT(m.id) AS meeting_count
            FROM workspaces w
            LEFT JOIN meetings m ON m.workspace_id = w.id
            GROUP BY w.id
            ORDER BY w.created_at DESC
        """)
        return [dict(r) for r in rows]


@app.delete("/workspaces/{workspace_id}")
async def delete_workspace(workspace_id: int):
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM workspaces WHERE id = $1", workspace_id
        )
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="Workspace not found")
        return {"ok": True}


@app.get("/workspaces/{workspace_id}/meetings")
async def list_workspace_meetings(workspace_id: int):
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, title, filename, date, summary FROM meetings WHERE workspace_id = $1 ORDER BY date DESC",
            workspace_id,
        )
        return [dict(r) for r in rows]


# ---- Meeting endpoints ----


@app.get("/meetings")
async def list_meetings(unorganized: bool = Query(default=False)):
    async with db_pool.acquire() as conn:
        if unorganized:
            rows = await conn.fetch(
                "SELECT id, title, filename, date, summary FROM meetings WHERE workspace_id IS NULL ORDER BY date DESC LIMIT 50"
            )
        else:
            rows = await conn.fetch(
                "SELECT id, title, filename, date, summary FROM meetings ORDER BY date DESC LIMIT 50"
            )
        return [dict(r) for r in rows]


@app.get("/meetings/{meeting_id}")
async def get_meeting(meeting_id: int):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, title, filename, date, transcript, summary, action_items, email_body FROM meetings WHERE id = $1",
            meeting_id,
        )
        if not row:
            raise HTTPException(status_code=404, detail="Meeting not found")
        result = dict(row)
        if isinstance(result.get("action_items"), str):
            result["action_items"] = json.loads(result["action_items"])
        return result


# ---- Analysis endpoint ----


async def extract_audio(input_path: str, output_path: str) -> None:
    proc = await asyncio.create_subprocess_exec(
        "ffmpeg", "-i", input_path, "-vn",
        "-acodec", "pcm_s16le", "-ar", "16000", "-ac", "1",
        output_path, "-y",
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    _, stderr = await proc.communicate()
    if proc.returncode != 0:
        raise RuntimeError(f"ffmpeg failed: {stderr.decode()[-500:]}")


async def transcribe(audio_path: str) -> str:
    logger.info("Sending audio to Whisper: %s (%.1f MB)", audio_path, os.path.getsize(audio_path) / (1024 * 1024))
    async with httpx.AsyncClient(timeout=3600.0) as client:
        with open(audio_path, "rb") as f:
            resp = await client.post(
                f"{WHISPER_URL}/asr",
                params={"language": "en", "output": "json", "encode": "true"},
                files={"audio_file": ("audio.wav", f, "audio/wav")},
            )
        logger.info("Whisper response: status=%d", resp.status_code)
        resp.raise_for_status()
        data = resp.json()
        return data.get("text", "")


async def _call_claude(prompt: str) -> str:
    """Call Claude via the Agent SDK (supports OAuth tokens)."""
    from claude_code_sdk import query as claude_query, ClaudeCodeOptions
    from claude_code_sdk.types import AssistantMessage, ResultMessage, TextBlock

    if not CLAUDE_API_KEY:
        raise RuntimeError("Claude API key not configured. Set CLAUDE_API_KEY.")

    env = {}
    if CLAUDE_API_KEY.startswith("sk-ant-oat"):
        env["CLAUDE_CODE_OAUTH_TOKEN"] = CLAUDE_API_KEY
    else:
        env["ANTHROPIC_API_KEY"] = CLAUDE_API_KEY

    options = ClaudeCodeOptions(
        max_turns=1,
        model="claude-sonnet-4-6",
        allowed_tools=[],
        env=env,
    )

    result_text = ""
    try:
        stream = claude_query(prompt=prompt, options=options)
        while True:
            try:
                message = await stream.__anext__()
            except StopAsyncIteration:
                break
            except Exception as iter_err:
                if "Unknown message type" in str(iter_err):
                    logger.debug("Skipping unknown SDK event: %s", iter_err)
                    continue
                raise

            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock):
                        result_text += block.text
            elif isinstance(message, ResultMessage):
                if message.is_error:
                    logger.error("Claude SDK error: %s", message.result)
                    raise RuntimeError(f"Claude SDK error: {message.result}")
                if message.result and not result_text:
                    result_text = message.result
    except RuntimeError:
        raise
    except Exception as e:
        if result_text:
            logger.warning("SDK error after getting text (ignored): %s", e)
            return result_text
        raise RuntimeError(f"Claude SDK error: {e}")

    return result_text or ""


# ---- Streaming TTS helpers ----

SENTENCE_END_RE = re.compile(r'(?<=[.!?])\s+')


class SentenceBuffer:
    """Buffer streaming tokens and emit complete sentences."""

    def __init__(self, min_length: int = 10):
        self.buffer = ""
        self.min_length = min_length

    def add_token(self, token: str) -> list[str]:
        self.buffer += token
        # Short-circuit: skip regex if no sentence-ending punctuation present
        if "." not in self.buffer and "!" not in self.buffer and "?" not in self.buffer:
            return []
        sentences = []
        while True:
            match = SENTENCE_END_RE.search(self.buffer)
            if not match:
                break
            sentence = self.buffer[:match.end()].strip()
            self.buffer = self.buffer[match.end():]
            if len(sentence) >= self.min_length:
                sentences.append(sentence)
        return sentences

    def flush(self) -> str | None:
        remaining = self.buffer.strip()
        self.buffer = ""
        return remaining if len(remaining) >= 3 else None


async def _stream_claude_tokens(prompt: str):
    """Async generator that yields text tokens from Claude as they arrive."""
    from claude_code_sdk import query as claude_query, ClaudeCodeOptions
    from claude_code_sdk.types import AssistantMessage, TextBlock

    if not CLAUDE_API_KEY:
        raise RuntimeError("Claude API key not configured. Set CLAUDE_API_KEY.")

    env = {}
    if CLAUDE_API_KEY.startswith("sk-ant-oat"):
        env["CLAUDE_CODE_OAUTH_TOKEN"] = CLAUDE_API_KEY
    else:
        env["ANTHROPIC_API_KEY"] = CLAUDE_API_KEY

    options = ClaudeCodeOptions(
        max_turns=1,
        model="claude-sonnet-4-6",
        allowed_tools=[],
        env=env,
    )

    stream = claude_query(prompt=prompt, options=options)
    while True:
        try:
            message = await stream.__anext__()
        except StopAsyncIteration:
            break
        except Exception as iter_err:
            if "Unknown message type" in str(iter_err):
                logger.debug("Skipping unknown SDK event: %s", iter_err)
                continue
            raise

        if isinstance(message, AssistantMessage):
            for block in message.content:
                if isinstance(block, TextBlock):
                    yield block.text


async def synthesize_and_send(sentence: str, ws: WebSocket):
    """Synthesize a sentence to PCM and send over WebSocket (non-blocking)."""
    import numpy as np

    def _synth():
        pcm_chunks = []
        for audio_chunk in piper_voice.synthesize(sentence):
            pcm_chunks.append((audio_chunk.audio_float_array * 32767).astype(np.int16).tobytes())
        return b"".join(pcm_chunks)

    pcm = await asyncio.to_thread(_synth)
    await ws.send_bytes(pcm)


async def analyze_with_claude(transcript: str) -> dict:
    if not CLAUDE_API_KEY:
        return {
            "title": "Untitled Meeting",
            "summary": "Claude API key not configured.",
            "action_items": [],
            "email_body": "",
        }

    prompt = ANALYSIS_PROMPT.format(transcript=transcript)
    text = await _call_claude(prompt)
    # Strip markdown fences if present (SDK often wraps output)
    stripped = text.strip()
    if stripped.startswith("```"):
        lines = stripped.split("\n")
        # Remove first line (```json or ```) and last line (```)
        lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        stripped = "\n".join(lines).strip()
    try:
        return json.loads(stripped)
    except (json.JSONDecodeError, ValueError):
        return {
            "title": "Untitled Meeting",
            "summary": stripped or "Analysis failed.",
            "action_items": [],
            "email_body": "",
        }


async def save_meeting(
    filename: str, transcript: str, analysis: dict, workspace_id: int | None = None
) -> int:
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """INSERT INTO meetings (title, filename, transcript, summary, action_items, email_body, workspace_id)
               VALUES ($1, $2, $3, $4, $5::jsonb, $6, $7)
               RETURNING id""",
            analysis.get("title", "Untitled Meeting"),
            filename,
            transcript,
            analysis.get("summary", ""),
            json.dumps(analysis.get("action_items", [])),
            analysis.get("email_body", ""),
            workspace_id,
        )
        return row["id"]


@app.post("/analyze")
async def analyze(
    file: UploadFile = File(...),
    workspace_id: int | None = Query(default=None),
):
    if not file.filename or not file.filename.lower().endswith((".mp4", ".m4a", ".mp3")):
        raise HTTPException(status_code=400, detail="Only MP4, M4A, and MP3 files are supported.")

    contents = await file.read()
    original_filename = file.filename
    file_size_mb = len(contents) / (1024 * 1024)
    logger.info("Upload received: %s (%.1f MB), workspace=%s", original_filename, file_size_mb, workspace_id)

    async def stream():
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, "input.mp4")
            audio_path = os.path.join(tmpdir, "audio.wav")

            yield json.dumps({"status": "Saving uploaded file..."}) + "\n"

            with open(input_path, "wb") as f:
                f.write(contents)

            yield json.dumps({"status": "Extracting audio with ffmpeg..."}) + "\n"
            try:
                await extract_audio(input_path, audio_path)
                audio_size_mb = os.path.getsize(audio_path) / (1024 * 1024)
                logger.info("Audio extracted: %.1f MB", audio_size_mb)
            except RuntimeError as e:
                logger.error("ffmpeg failed for %s: %s", original_filename, e)
                yield json.dumps({"error": str(e)}) + "\n"
                return

            yield json.dumps({"status": "Transcribing with Whisper (this may take a while)..."}) + "\n"
            try:
                import time as _time
                t0 = _time.monotonic()
                transcript = await transcribe(audio_path)
                elapsed = _time.monotonic() - t0
                logger.info("Transcription done in %.1fs (%d chars)", elapsed, len(transcript))
            except Exception as e:
                logger.error("Whisper transcription failed for %s: %s", original_filename, e)
                yield json.dumps({"error": f"Whisper transcription failed: {e}"}) + "\n"
                return

            if not transcript.strip():
                logger.warning("No speech detected in %s", original_filename)
                yield json.dumps({"error": "No speech detected in the audio."}) + "\n"
                return

            yield json.dumps({"status": "Analyzing with Claude..."}) + "\n"
            try:
                analysis = await analyze_with_claude(transcript)
                logger.info("Claude analysis complete for %s", original_filename)
            except Exception as e:
                logger.error("Claude analysis failed for %s: %s", original_filename, e)
                yield json.dumps({"error": f"Claude analysis failed: {e}"}) + "\n"
                return

            try:
                meeting_id = await save_meeting(
                    original_filename, transcript, analysis, workspace_id
                )
                logger.info("Meeting saved: id=%s, file=%s", meeting_id, original_filename)
            except Exception as e:
                logger.error("Save failed for %s: %s", original_filename, e)
                yield json.dumps({"error": f"Failed to save meeting: {e}"}) + "\n"
                return

            result = {
                "id": meeting_id,
                "transcript": transcript,
                "summary": analysis.get("summary", ""),
                "action_items": analysis.get("action_items", []),
                "email_body": analysis.get("email_body", ""),
            }
            yield json.dumps({"result": result}) + "\n"

    return StreamingResponse(
        stream(),
        media_type="application/x-ndjson",
        headers={"X-Accel-Buffering": "no"},  # prevent nginx from buffering the stream
    )


# ---- Document helpers ----


def _detect_mime(filename: str) -> str:
    mime, _ = mimetypes.guess_type(filename)
    return mime or "application/octet-stream"


def _extract_text_sync(data: bytes, mime_type: str, filename: str) -> str | None:
    """Extract text from document bytes. Returns None if unsupported."""
    ext = os.path.splitext(filename.lower())[1]
    try:
        if ext == ".pdf" or mime_type == "application/pdf":
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    parts.append(text)
            return "\n".join(parts) if parts else ""
        elif ext == ".docx" or mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs if p.text)
        elif ext in (".xlsx", ".xls") or mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows.append("\t".join(cells))
            return "\n".join(rows)
        elif ext in (".txt", ".md", ".csv", ".log", ".json", ".yaml", ".yml") or (
            mime_type and mime_type.startswith("text/")
        ):
            return data.decode("utf-8", errors="replace")
        else:
            return None
    except Exception as e:
        logger.warning("Text extraction failed for %s: %s", filename, e)
        return None


async def _extract_and_store(doc_id: int, data: bytes, mime_type: str, filename: str):
    """Background task: extract text and update documents row."""
    try:
        extracted = await asyncio.to_thread(_extract_text_sync, data, mime_type, filename)
        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE documents SET extracted_text = $1 WHERE id = $2",
                extracted,
                doc_id,
            )
        logger.info("Text extraction complete for doc %d: %s chars", doc_id, len(extracted) if extracted else 0)
    except Exception as e:
        logger.error("Text extraction error for doc %d: %s", doc_id, e)


def _get_minio_client():
    if minio_client is None:
        raise HTTPException(status_code=503, detail="Document storage not available.")
    return minio_client.create_client(
        "s3",
        endpoint_url=f"http://{MINIO_ENDPOINT}",
        aws_access_key_id=MINIO_ACCESS_KEY,
        aws_secret_access_key=MINIO_SECRET_KEY,
        region_name="us-east-1",
    )


# ---- Document endpoints ----


@app.post("/workspaces/{workspace_id}/documents")
async def upload_document(
    workspace_id: int,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
):
    async with db_pool.acquire() as conn:
        ws = await conn.fetchrow("SELECT id FROM workspaces WHERE id = $1", workspace_id)
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")

    data = await file.read()
    filename = file.filename or "upload"
    mime_type = _detect_mime(filename)
    object_key = f"workspaces/{workspace_id}/documents/{uuid.uuid4()}_{filename}"

    async with _get_minio_client() as client:
        await client.put_object(
            Bucket=MINIO_BUCKET,
            Key=object_key,
            Body=data,
            ContentType=mime_type,
        )

    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """INSERT INTO documents (workspace_id, filename, object_key, file_size, mime_type)
               VALUES ($1, $2, $3, $4, $5)
               RETURNING id, filename, file_size, mime_type, uploaded_at""",
            workspace_id, filename, object_key, len(data), mime_type,
        )
    doc = dict(row)
    doc["has_text"] = False

    background_tasks.add_task(_extract_and_store, doc["id"], data, mime_type, filename)

    return doc


@app.get("/workspaces/{workspace_id}/documents")
async def list_documents(workspace_id: int):
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT id, filename, file_size, mime_type, uploaded_at,
                      (extracted_text IS NOT NULL) AS has_text,
                      (extracted_text IS NULL) AS pending
               FROM documents
               WHERE workspace_id = $1
               ORDER BY uploaded_at DESC""",
            workspace_id,
        )
    return [dict(r) for r in rows]


@app.get("/workspaces/{workspace_id}/documents/{doc_id}/download")
async def download_document(workspace_id: int, doc_id: int):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT filename, object_key, mime_type FROM documents WHERE id = $1 AND workspace_id = $2",
            doc_id, workspace_id,
        )
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")

    async with _get_minio_client() as client:
        response = await client.get_object(Bucket=MINIO_BUCKET, Key=row["object_key"])
        data = await response["Body"].read()

    return StreamingResponse(
        iter([data]),
        media_type=row["mime_type"],
        headers={"Content-Disposition": f'attachment; filename="{row["filename"]}"'},
    )


@app.delete("/workspaces/{workspace_id}/documents/{doc_id}")
async def delete_document(workspace_id: int, doc_id: int):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT object_key FROM documents WHERE id = $1 AND workspace_id = $2",
            doc_id, workspace_id,
        )
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")

    try:
        async with _get_minio_client() as client:
            await client.delete_object(Bucket=MINIO_BUCKET, Key=row["object_key"])
    except Exception as e:
        logger.warning("MinIO delete failed for %s: %s", row["object_key"], e)

    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM documents WHERE id = $1", doc_id)

    return {"ok": True}


@app.get("/workspaces/{workspace_id}/documents/{doc_id}/text")
async def get_document_text(workspace_id: int, doc_id: int):
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT filename, extracted_text FROM documents WHERE id = $1 AND workspace_id = $2",
            doc_id, workspace_id,
        )
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"filename": row["filename"], "text": row["extracted_text"]}


# ---- Chat endpoint ----


def _esc_xml(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


async def _build_chat_prompt(
    meeting_ids: list[int],
    question: str,
    include_transcripts: list[int] | None = None,
    include_document_ids: list[int] | None = None,
) -> tuple[str, bool]:
    """Build a chat prompt from meeting data and optional documents. Returns (prompt, warned)."""
    include_transcripts = include_transcripts or []
    include_document_ids = include_document_ids or []

    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, title, date, summary, action_items, transcript FROM meetings WHERE id = ANY($1::int[])",
            meeting_ids,
        )
    if not rows:
        raise HTTPException(status_code=404, detail="No meetings found.")

    meetings = sorted([dict(r) for r in rows], key=lambda m: m["date"], reverse=True)
    context_parts = []
    total_chars = 0
    max_chars = 150_000
    warned = False

    for m in meetings:
        items = m["action_items"]
        if isinstance(items, str):
            items = json.loads(items)

        part = f'<meeting id="{m["id"]}" title="{_esc_xml(m["title"])}" date="{m["date"]}">\n'
        part += f"<summary>{_esc_xml(m['summary'] or '')}</summary>\n"
        if items:
            part += "<action_items>\n" + "\n".join(f"- {item}" for item in items) + "\n</action_items>\n"

        if m["id"] in include_transcripts and m["transcript"]:
            transcript_addition = f"<transcript>{m['transcript']}</transcript>\n"
            if total_chars + len(part) + len(transcript_addition) > max_chars:
                warned = True
            else:
                part += transcript_addition

        part += "</meeting>"
        context_parts.append(part)
        total_chars += len(part)

    context = "\n\n".join(context_parts)

    # Add document context if requested
    doc_parts = []
    if include_document_ids:
        async with db_pool.acquire() as conn:
            doc_rows = await conn.fetch(
                "SELECT id, filename, extracted_text FROM documents WHERE id = ANY($1::int[]) AND extracted_text IS NOT NULL",
                include_document_ids,
            )
        for dr in doc_rows:
            text = dr["extracted_text"] or ""
            addition = f'<document filename="{_esc_xml(dr["filename"])}">\n{text}\n</document>'
            if total_chars + len(addition) > max_chars:
                warned = True
                break
            doc_parts.append(addition)
            total_chars += len(addition)

    doc_context = "\n\n".join(doc_parts)

    prompt = (
        f"{CHAT_SYSTEM_PROMPT}\n\n"
        f"<meetings>\n{context}\n</meetings>\n\n"
    )
    if doc_context:
        prompt += f"<documents>\n{doc_context}\n</documents>\n\n"
    prompt += f"Question: {question}"

    return prompt, warned


@app.post("/chat")
async def chat(body: ChatRequest):
    if not CLAUDE_API_KEY:
        raise HTTPException(status_code=500, detail="Claude API key not configured.")

    if not body.meeting_ids or not body.question.strip():
        raise HTTPException(status_code=400, detail="meeting_ids and question are required.")

    prompt, warned = await _build_chat_prompt(
        body.meeting_ids, body.question, body.include_transcripts, body.include_document_ids
    )

    async def stream():
        if warned:
            yield json.dumps({"warning": "Context exceeded limit. Some transcripts were excluded."}) + "\n"

        try:
            response_text = await _call_claude(prompt)
            yield json.dumps({"token": response_text}) + "\n"
        except Exception as e:
            yield json.dumps({"error": str(e)}) + "\n"
            return

        yield json.dumps({"done": True}) + "\n"

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# ---- Analyze text endpoint (for live transcriptions) ----


@app.post("/analyze-text")
async def analyze_text(body: AnalyzeTextRequest):
    text = body.text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="Text is required.")

    async def stream():
        yield json.dumps({"status": "Analyzing with Claude..."}) + "\n"
        try:
            analysis = await analyze_with_claude(text)
        except Exception as e:
            logger.error("Claude analysis failed for text input: %s", e)
            yield json.dumps({"error": f"Claude analysis failed: {e}"}) + "\n"
            return

        try:
            meeting_id = await save_meeting(
                "live-transcription", text, analysis, body.workspace_id
            )
        except Exception as e:
            logger.error("Save failed for text input: %s", e)
            yield json.dumps({"error": f"Failed to save meeting: {e}"}) + "\n"
            return

        result = {
            "id": meeting_id,
            "transcript": text,
            "summary": analysis.get("summary", ""),
            "action_items": analysis.get("action_items", []),
            "email_body": analysis.get("email_body", ""),
        }
        yield json.dumps({"result": result}) + "\n"

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# ---- Streaming TTS WebSocket ----


@app.websocket("/ws/tts-stream")
async def ws_tts_stream(ws: WebSocket):
    await ws.accept()
    try:
        if not piper_voice:
            await ws.send_json({"type": "error", "message": "Streaming TTS not available — Piper model not loaded"})
            await ws.close()
            return

        request = await ws.receive_json()
        question = request.get("question", "").strip()
        meeting_ids = request.get("meeting_ids", [])
        include_transcripts = request.get("include_transcripts", [])
        include_document_ids = request.get("include_document_ids", [])

        if not question or not meeting_ids:
            await ws.send_json({"type": "error", "message": "question and meeting_ids required"})
            await ws.close()
            return

        # Build prompt from meeting context
        prompt, warned = await _build_chat_prompt(
            meeting_ids, question, include_transcripts, include_document_ids
        )

        if warned:
            await ws.send_json({"type": "warning", "message": "Context exceeded limit. Some transcripts were excluded."})

        # Send audio config
        await ws.send_json({
            "type": "audio_config",
            "sample_rate": piper_voice.config.sample_rate,
            "channels": 1,
            "bit_depth": 16,
            "encoding": "pcm_s16le",
        })

        sentence_buffer = SentenceBuffer()
        text_parts = []

        async for token in _stream_claude_tokens(prompt):
            text_parts.append(token)
            await ws.send_json({"type": "token", "text": token})

            for sentence in sentence_buffer.add_token(token):
                await ws.send_json({"type": "sentence", "text": sentence})
                await synthesize_and_send(sentence, ws)

        # Flush remaining text
        remaining = sentence_buffer.flush()
        if remaining:
            await ws.send_json({"type": "sentence", "text": remaining})
            await synthesize_and_send(remaining, ws)

        await ws.send_json({"type": "done", "full_text": "".join(text_parts)})

    except WebSocketDisconnect:
        logger.info("TTS stream client disconnected")
    except Exception as e:
        logger.error("TTS stream error: %s", e)
        try:
            await ws.send_json({"type": "error", "message": str(e)})
        except Exception:
            pass


# ---- LiveKit integration ----

import time as _time


class LiveKitTokenRequest(BaseModel):
    room: str = "meeting"
    identity: str | None = None


def _livekit_token(claims: dict) -> str:
    return jwt.encode(claims, LIVEKIT_API_SECRET, algorithm="HS256")


@app.post("/livekit/token")
async def livekit_token(body: LiveKitTokenRequest):
    """Generate a LiveKit JWT for a subscriber joining a room."""
    identity = body.identity or f"listener-{uuid.uuid4().hex[:8]}"
    now = int(_time.time())
    claims = {
        "video": {
            "roomJoin": True,
            "room": body.room,
            "canPublish": False,
            "canSubscribe": True,
        },
        "sub": identity,
        "iss": LIVEKIT_API_KEY,
        "exp": now + 3600,
        "nbf": now,
        "jti": uuid.uuid4().hex,
    }
    token = _livekit_token(claims)
    return {"token": token, "url": LIVEKIT_WS_URL, "identity": identity}


@app.get("/livekit/rooms")
async def livekit_rooms():
    """List active LiveKit rooms using the internal admin API."""
    now = int(_time.time())
    claims = {
        "video": {"roomList": True},
        "sub": "server",
        "iss": LIVEKIT_API_KEY,
        "exp": now + 60,
        "nbf": now,
        "jti": uuid.uuid4().hex,
    }
    admin_token = _livekit_token(claims)
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                f"{LIVEKIT_INTERNAL_URL}/twirp/livekit.RoomService/ListRooms",
                headers={
                    "Authorization": f"Bearer {admin_token}",
                    "Content-Type": "application/json",
                },
                json={},
                timeout=5.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                return {"rooms": [r.get("name") for r in data.get("rooms", [])]}
    except Exception as e:
        logger.warning("Failed to list LiveKit rooms: %s", e)
    return {"rooms": []}


# ---- WebSocket proxy to WhisperLive ----


@app.websocket("/ws/transcribe")
async def ws_transcribe(ws: WebSocket):
    await ws.accept()
    whisper_ws = None

    try:
        # Wait for start message from browser
        init_data = await ws.receive_json()
        if init_data.get("action") != "start":
            await ws.send_json({"error": "Expected {action: 'start'}"})
            await ws.close()
            return

        language = init_data.get("language", "en")
        uid = str(uuid.uuid4())

        # Connect to WhisperLive
        whisper_ws = await websockets.connect(WHISPER_LIVE_URL)

        # Send config to WhisperLive
        config = {
            "uid": uid,
            "language": language,
            "model": "base",
            "task": "transcribe",
            "use_vad": True,
        }
        await whisper_ws.send(json.dumps(config))

        # Relay loop: browser audio -> WhisperLive, WhisperLive segments -> browser
        async def relay_from_whisper():
            try:
                async for message in whisper_ws:
                    if isinstance(message, str):
                        data = json.loads(message)
                        await ws.send_json(data)
            except websockets.ConnectionClosed:
                pass
            except WebSocketDisconnect:
                pass

        relay_task = asyncio.create_task(relay_from_whisper())

        try:
            while True:
                message = await ws.receive()

                if message.get("type") == "websocket.disconnect":
                    break

                if "text" in message:
                    data = json.loads(message["text"])
                    if data.get("action") == "stop":
                        await whisper_ws.send(b"END_OF_AUDIO")
                        await asyncio.sleep(0.5)
                        await ws.send_json({"done": True})
                        break

                if "bytes" in message and message["bytes"]:
                    # Browser sends float32 PCM audio bytes, forward directly
                    await whisper_ws.send(message["bytes"])

        except WebSocketDisconnect:
            pass
        finally:
            relay_task.cancel()
            try:
                await relay_task
            except asyncio.CancelledError:
                pass

    except Exception as e:
        logger.error("WebSocket transcribe error: %s", e)
        try:
            await ws.send_json({"error": str(e)})
        except Exception:
            pass
    finally:
        if whisper_ws:
            await whisper_ws.close()
